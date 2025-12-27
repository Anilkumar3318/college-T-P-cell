import customtkinter as ctk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from utils import (
    batch_clear_widgets, safe_int_convert, safe_float_convert,
    extract_numeric_value, create_autopct_function, COLORS, CHART_COLORS,
    resource_path, get_cached_students, get_cached_companies, get_cached_placements,
    get_matplotlib, create_optimized_figure, embed_chart_in_frame, invalidate_cache,
    performance_monitor
)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class TP_Manager:
    def __init__(self, root):
        self.root = root
        self.username = None
        self.root.title("TP_Manager-Training and Placement Management System")

        # Set fullscreen
        self.root.overrideredirect(True)
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}+0+0")

        self.is_maximized = True
        self.normal_geometry = None

        from student import StudentManager
        from company import CompanyManager
        from placed_student import PlacedStudentManager

        self.student_manager = StudentManager()
        self.company_manager = CompanyManager()
        self.placed_student_manager = PlacedStudentManager()

        self.current_section = None
        self.current_action = None

        self._preload_data()
        self.show_login_screen()

    def _preload_data(self):
        import threading
        
        def preload():
            try:
                if self.student_manager.collection is not None:
                    get_cached_students(self.student_manager.collection, limit=100)
                if self.company_manager.collection is not None:
                    get_cached_companies(self.company_manager.collection, limit=50)
                if self.placed_student_manager.collection is not None:
                    get_cached_placements(self.placed_student_manager.collection, limit=50)
                print("✅ Data preloading completed")
            except Exception as e:
                print(f"⚠️ Data preloading failed: {e}")
        
        threading.Thread(target=preload, daemon=True).start()

    def show_login_screen(self):
        from login import LoginFrame
        self.login_frame = LoginFrame(self.root, self.on_login_success)

    def on_login_success(self, username):
        self.username = username
        self.root.title(f"TP_Manager - Training and Placement Management System - Logged in as: {username}")
        self.setup_gui()

    def setup_gui(self):
        # --- CUSTOM TITLE BAR ---
        self.title_bar = ctk.CTkFrame(self.root, height=30, fg_color="#2b2b2b", corner_radius=0)
        self.title_bar.pack(fill='x', side='top')
        self.title_bar.pack_propagate(False)

        # Title bar content
        title_container = ctk.CTkFrame(self.title_bar, fg_color="transparent")
        title_container.pack(fill='both', padx=10)

        app_title = ctk.CTkLabel(title_container, text="TP_Manager - Training and Placement Management System",
                                 font=("Arial", 12, "bold"), text_color="white")
        app_title.pack(side='left')

        # Window controls
        controls_frame = ctk.CTkFrame(title_container, fg_color="transparent")
        controls_frame.pack(side='right')

        button_style = {
            "width": 30,
            "height": 20,
            "corner_radius": 3,
            "font": ("Arial", 10, "bold")
        }

        minimize_button = ctk.CTkButton(controls_frame, text="—", command=self.minimize_window,
                                        fg_color="transparent", hover_color="#404040", **button_style)
        minimize_button.pack(side='left', padx=2)

        self.maximize_button = ctk.CTkButton(controls_frame, text="❐", command=self.toggle_maximize,
                                             fg_color="transparent", hover_color="#404040", **button_style)
        self.maximize_button.pack(side='left', padx=2)

        close_button = ctk.CTkButton(controls_frame, text="×", command=self.close_window,
                                     fg_color="transparent", hover_color="#E74C3C", **button_style)
        close_button.pack(side='left', padx=2)

        # --- MAIN CONTENT FRAME ---
        self.main_frame = ctk.CTkFrame(self.root, fg_color=COLORS["content_frame"])
        self.main_frame.pack(fill='both', expand=True, padx=0, pady=0)

        # --- TOP FRAME ---
        self.setup_top_frame()

        # --- NAVIGATION FRAME ---
        self.setup_navigation_frame()

        # --- CONTENT FRAME ---
        self.setup_content_frame()

        # Show home page by default and select home button
        self.select_section('home')

    def setup_top_frame(self):
        # Keep frame height same as before
        frame_height = max(140, int(self.root.winfo_screenheight() * 0.12))  # At least 140px or 12% of screen height
        # Use white background to match logo
        self.page_option_frame = ctk.CTkFrame(master=self.main_frame, height=frame_height, corner_radius=10,
                                              fg_color="#FFFFFF")
        self.page_option_frame.pack(side='top', fill='x', padx=8, pady=3)
        self.page_option_frame.pack_propagate(False)

        # Container frame for logo + title (centered together)
        header_container = ctk.CTkFrame(self.page_option_frame, fg_color="#FFFFFF", corner_radius=0)
        header_container.pack(expand=True, fill='both', padx=5, pady=2)

        # Inner frame to hold logo and title side by side, centered
        inner_frame = ctk.CTkFrame(header_container, fg_color="#FFFFFF", corner_radius=0)
        inner_frame.pack(expand=True, pady=2)

        # Logo next to title (original size 130x130)
        try:
            img = Image.open(resource_path("mbs_logo.png"))
            img = img.resize((130, 130), Image.Resampling.LANCZOS)
            ctk_image = ctk.CTkImage(light_image=img, dark_image=img, size=(130, 130))
            self.logo_label = ctk.CTkLabel(inner_frame, image=ctk_image, text="", 
                                           width=130, height=130, fg_color="#FFFFFF")
            self.logo_label.pack(side='left', padx=0, pady=0)
        except (FileNotFoundError, Exception) as e:
            print(f"Logo loading failed: {e}")
            self.logo_label = ctk.CTkLabel(inner_frame, text="🎓", 
                                           width=130, height=130,
                                           font=("Arial", 50), text_color="black", fg_color="#FFFFFF")
            self.logo_label.pack(side='left', padx=0, pady=0)

        # Title next to logo
        title_label = ctk.CTkLabel(
            inner_frame,
            text="MAHANT BACHITTAR SINGH COLLEGE OF ENGINEERING & TECHNOLOGY\nTRAINING AND PLACEMENT CELL",
            font=("Arial", 22, "bold"),
            text_color="black",
            justify="center",
            fg_color="#FFFFFF"
        )
        title_label.pack(side='left', padx=0, pady=0)

    def setup_navigation_frame(self):
        self.page_option_frame_1 = ctk.CTkFrame(master=self.main_frame, height=60, corner_radius=10,
                                                fg_color=COLORS["nav_frame"])
        self.page_option_frame_1.pack(side='top', fill='x', padx=2, pady=2)

        # Navigation buttons with enhanced interactive styling
        nav_button_style = {
            "text_color": "white",
            "font": ("Arial", 14, "normal"),  # Start with normal weight
            "width": 140,
            "height": 40,
            "corner_radius": 8,
            "fg_color": COLORS["button_default"],
            "hover_color": COLORS["hover"],  # Nice hover effect
            "border_width": 2,
            "border_color": COLORS["nav_frame"]
        }

        self.nav_buttons = {}

        self.nav_buttons['home'] = ctk.CTkButton(master=self.page_option_frame_1, text='🏠 HOME',
                                                 command=lambda: self.select_section('home'), **nav_button_style)
        self.nav_buttons['home'].place(x=20, y=10)

        self.nav_buttons['student'] = ctk.CTkButton(master=self.page_option_frame_1, text='📚 STUDENT',
                                                    command=lambda: self.select_section('student'),
                                                    **nav_button_style)
        self.nav_buttons['student'].place(x=180, y=10)

        self.nav_buttons['company'] = ctk.CTkButton(master=self.page_option_frame_1, text='🏢 COMPANY',
                                                    command=lambda: self.select_section('company'),
                                                    **nav_button_style)
        self.nav_buttons['company'].place(x=340, y=10)

        self.nav_buttons['placed'] = ctk.CTkButton(master=self.page_option_frame_1, text='🎓 PLACED STUDENT',
                                                   command=lambda: self.select_section('placed'), **nav_button_style)
        self.nav_buttons['placed'].place(x=500, y=10)

        # Database button on the right side
        db_button_style = {
            "text_color": "white",
            "font": ("Arial", 12, "bold"),
            "width": 120,
            "height": 40,
            "corner_radius": 8,
            "fg_color": "#E74C3C",
            "hover_color": "#C0392B",
            "border_width": 2,
            "border_color": "#E74C3C"
        }

        self.nav_buttons['database'] = ctk.CTkButton(master=self.page_option_frame_1, text='🗄️ DATABASE',
                                                     command=self.open_database_settings, **db_button_style)
        self.nav_buttons['database'].place(x=1200, y=10)

        # Sub-navigation frame (for action buttons)
        self.sub_nav_frame = ctk.CTkFrame(master=self.main_frame, height=50, corner_radius=10,
                                          fg_color=COLORS["section_frame"])
        self.sub_nav_frame.pack(side='top', fill='x', padx=8, pady=3)
        self.sub_nav_frame.pack_forget()  # Hidden by default

    def select_section(self, section):
        """Handle main navigation selection with enhanced interactive button behavior"""
        self.current_section = section
        self.current_action = None

        # Safety check - ensure nav_buttons exist
        if not hasattr(self, 'nav_buttons') or not self.nav_buttons:
            print(f"Navigation not ready yet, section set to: {section}")
            return

        # Update button styles with enhanced interactivity
        for key, btn in self.nav_buttons.items():
            if key == 'database':  # Skip database button styling
                continue
                
            if key == section:
                # Selected button: Enhanced blue styling with animation-like effect
                btn.configure(
                    fg_color=COLORS["selected"],  # Blue background
                    hover_color="#1a5490",  # Darker blue on hover
                    text_color="white",
                    border_width=3,
                    border_color="#ffffff",  # White border for selected
                    font=("Arial", 14, "bold")  # Keep bold font
                )
            else:
                # Unselected buttons: Enhanced default styling
                btn.configure(
                    fg_color=COLORS["button_default"],  # Default gray
                    hover_color=COLORS["hover"],  # Light blue on hover
                    text_color="white",
                    border_width=2,
                    border_color=COLORS["nav_frame"],  # Subtle border
                    font=("Arial", 14, "normal")  # Normal font weight
                )

        # Show appropriate sub-navigation
        if section == 'home':
            self.sub_nav_frame.pack_forget()
            self.show_home()
        elif section == 'student':
            # Show sub-navigation for student
            self.show_sub_navigation(section)
            # Open charts by default
            self.student_manager.show_chart_interface(self.content_area)
        elif section == 'company':
            # Show sub-navigation for company
            self.show_sub_navigation(section)
            # Open charts by default
            self.company_manager.show_chart_interface(self.content_area)
        elif section == 'placed':
            # Show sub-navigation for placed students
            self.show_sub_navigation(section)
            # Open charts by default
            self.placed_student_manager.show_chart_interface(self.content_area)
        else:
            self.show_sub_navigation(section)

    def show_sub_navigation(self, section):
        """Show action buttons for selected section"""
        # Safety check - ensure sub_nav_frame exists
        if not hasattr(self, 'sub_nav_frame'):
            print(f"Sub-navigation not ready yet for section: {section}")
            return
            
        # Clear previous sub-navigation
        for widget in self.sub_nav_frame.winfo_children():
            widget.destroy()

        self.sub_nav_frame.pack(side='top', fill='x', padx=10, pady=5)

        ctk.CTkLabel(self.sub_nav_frame, text=f"{section.upper()} ACTIONS:",
                     font=("Arial", 12, "bold"), text_color="white").pack(side='left', padx=20)

        action_button_style = {
            "text_color": "white",
            "font": ("Arial", 12, "normal"),  # Start with normal weight
            "width": 100,
            "height": 30,
            "corner_radius": 6,
            "fg_color": COLORS["button_default"],
            "hover_color": COLORS["hover"],  # Enhanced hover effect
            "border_width": 2,
            "border_color": COLORS["section_frame"]
        }

        self.action_buttons = {}

        actions = [
            ('add', '➕ ADD', COLORS["success"]),
            ('view', '👁 VIEW', COLORS["info"]),
            ('edit', '✏️ EDIT', COLORS["warning"]),
            ('delete', '🗑️ DELETE', COLORS["error"]),
            ('chart', '📊 CHART', COLORS["info"])
        ]

        for action_key, action_text, color in actions:
            self.action_buttons[action_key] = ctk.CTkButton(
                self.sub_nav_frame,
                text=action_text,
                command=lambda a=action_key: self.select_action(a),
                **action_button_style
            )
            self.action_buttons[action_key].pack(side='left', padx=5)

        # Clear content area
        self.clear_content()

    def select_action(self, action):
        """Handle action selection with enhanced interactive button behavior"""
        self.current_action = action

        # Update action button styles with enhanced interactivity
        for key, btn in self.action_buttons.items():
            if key == action:
                # Selected action: Enhanced blue styling
                btn.configure(
                    fg_color=COLORS["selected"],  # Blue background
                    hover_color="#1a5490",  # Darker blue on hover
                    border_color="#ffffff",  # White border
                    border_width=3,
                    text_color="white",
                    font=("Arial", 12, "bold")  # Bold font for selected
                )
            else:
                # Unselected actions: Enhanced default style
                btn.configure(
                    fg_color=COLORS["button_default"],  # Default gray
                    hover_color=COLORS["hover"],  # Light blue on hover
                    border_color=COLORS["section_frame"],  # Subtle border
                    border_width=2,
                    text_color="white",
                    font=("Arial", 12, "normal")  # Normal font weight
                )

        # Route to appropriate function
        if self.current_section == 'student':
            if action == 'add':
                self.student_manager.show_add_interface(self.content_area)
            elif action == 'view':
                self.student_manager.show_view_interface(self.content_area)
            elif action == 'edit':
                self.student_manager.show_edit_interface(self.content_area)
            elif action == 'delete':
                self.student_manager.show_delete_interface(self.content_area)
            elif action == 'chart':
                self.student_manager.show_chart_interface(self.content_area)

        elif self.current_section == 'company':
            if action == 'add':
                self.company_manager.show_add_interface(self.content_area)
            elif action == 'view':
                self.company_manager.show_view_interface(self.content_area)
            elif action == 'edit':
                self.company_manager.show_edit_interface(self.content_area)
            elif action == 'delete':
                self.company_manager.show_delete_interface(self.content_area)
            elif action == 'chart':
                self.company_manager.show_chart_interface(self.content_area)

        elif self.current_section == 'placed':
            if action == 'add':
                self.placed_student_manager.show_add_interface(self.content_area)
            elif action == 'view':
                self.placed_student_manager.show_view_interface(self.content_area)
            elif action == 'edit':
                self.placed_student_manager.show_edit_interface(self.content_area)
            elif action == 'delete':
                self.placed_student_manager.show_delete_interface(self.content_area)
            elif action == 'chart':
                self.placed_student_manager.show_chart_interface(self.content_area)

    def setup_content_frame(self):
        self.content_area = ctk.CTkScrollableFrame(self.main_frame, fg_color=COLORS["content_frame"])
        self.content_area.pack(side='top', fill='both', expand=True, padx=30, pady=30)

    def clear_content(self):
        """Efficiently clear content area"""
        # Safety check - ensure content_area exists
        if hasattr(self, 'content_area') and self.content_area:
            batch_clear_widgets(self.content_area)

    @performance_monitor
    def show_home(self):
        # Safety check - ensure content_area exists
        if not hasattr(self, 'content_area'):
            return
            
        self.clear_content()

        # Header frame with title and export button
        header_frame = ctk.CTkFrame(self.content_area, fg_color=COLORS["content_frame"])
        header_frame.pack(fill='x', padx=5, pady=5)

        # Title (smaller)
        title = ctk.CTkLabel(header_frame, text="📊 ANALYTICS DASHBOARD",
                             font=("Arial", 20, "bold"), text_color=COLORS["info"])
        title.pack(side='left', padx=10, pady=8)

        # Export button (on the right)
        export_btn = ctk.CTkButton(header_frame, text="📊 EXPORT TO EXCEL",
                                   command=self.export_home_charts_to_excel,
                                   height=32, width=160, font=("Arial", 11, "bold"),
                                   fg_color=COLORS["success"])
        export_btn.pack(side='right', padx=10, pady=8)

        # Create scrollable frame for charts - make it scrollable
        charts_container = ctk.CTkFrame(self.content_area, fg_color=COLORS["content_frame"])
        charts_container.pack(fill='both', expand=True, padx=5, pady=5)

        # Loading indicator
        loading_label = ctk.CTkLabel(charts_container, text="⏳ Loading analytics...",
                                     font=("Arial", 14), text_color=COLORS["info"])
        loading_label.pack(pady=50)
        charts_container.update()

        try:
            # Get data from databases using cached functions for better performance
            students = get_cached_students(self.student_manager.collection, limit=500)
            companies = get_cached_companies(self.company_manager.collection, limit=300)
            placements = get_cached_placements(self.placed_student_manager.collection, limit=300)

            # Remove loading label
            loading_label.destroy()

            # Lazy load matplotlib only when needed
            plt, FigureCanvasTkAgg, Figure = get_matplotlib()
            import numpy as np

            # Row 1: 3 charts side by side
            row1_frame = ctk.CTkFrame(charts_container, fg_color=COLORS["content_frame"])
            row1_frame.pack(fill='both', expand=True, padx=5, pady=5)

            # Chart 1: Students by Branch (Pie Chart)
            self.create_chart_frame(row1_frame, "Students by Branch",
                                    self.get_students_by_branch(students), "pie", row1_frame)

            # Chart 2: Company Packages Distribution (Bar Chart with ranges)
            pkg_dist = self.get_package_distribution(companies)
            self.create_chart_frame(row1_frame, "Company Package Ranges (LPA)",
                                    pkg_dist, "bar_with_range", row1_frame)

            # Chart 3: Total Records Count (Bar Chart with range)
            records_data = self.get_records_count(students, companies, placements)
            self.create_chart_frame(row1_frame, "Total Records Count",
                                    (records_data[0], records_data[1], records_data[2]), "bar_with_range", row1_frame)

            # Row 2: 1 chart (full width)
            row2_frame = ctk.CTkFrame(charts_container, fg_color=COLORS["content_frame"])
            row2_frame.pack(fill='both', expand=True, padx=5, pady=5)

            # Chart 4: Students vs Placed by Branch (Grouped Bar Chart)
            students_vs_placed = self.get_students_vs_placed_by_branch(students, placements)
            self.create_chart_frame(row2_frame, "Students vs Placed by Branch",
                                    students_vs_placed, "grouped_bar", row2_frame)

        except Exception as e:
            loading_label.destroy()
            error_label = ctk.CTkLabel(charts_container, text=f"Error loading analytics: {e}",
                                       font=("Arial", 14), text_color=COLORS["error"])
            error_label.pack(pady=50)

    def create_chart_frame(self, parent, title, data, chart_type, row_frame):
        """Create a chart frame with matplotlib - optimized version"""
        # Lazy load matplotlib
        plt, FigureCanvasTkAgg, Figure = get_matplotlib()

        frame = ctk.CTkFrame(row_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        frame.pack(side='left', fill='both', expand=True, padx=5, pady=5)

        # Title
        title_label = ctk.CTkLabel(frame, text=title, font=("Arial", 14, "bold"),
                                   text_color=COLORS["info"])
        title_label.pack(pady=10)

        try:
            # Create figure with optimized settings for better performance
            fig = Figure(figsize=(6.5, 5), dpi=80, facecolor='#2b2b2b', edgecolor='none')  # Reduced DPI
            ax = fig.add_subplot(111)
            ax.set_facecolor('#1e1e1e')
            fig.subplots_adjust(left=0.12, right=0.95, top=0.88, bottom=0.18)

            # Plot based on type with optimized rendering
            if chart_type == "pie" and data:
                labels, sizes = data
                # Limit to top 8 items for better performance
                if len(labels) > 8:
                    labels, sizes = labels[:8], sizes[:8]
                
                # Use optimized autopct function from utils
                autopct_func = create_autopct_function(sizes)
                ax.pie(sizes, labels=labels, autopct=autopct_func, colors=CHART_COLORS["primary"][:len(labels)],
                       textprops={'color': 'white', 'fontsize': 9, 'weight': 'bold'})
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)

            elif chart_type == "bar" and data:
                labels, values = data
                bars = ax.bar(range(len(labels)), values, color=CHART_COLORS["primary"][:len(labels)])

                # Add optimized value labels
                for i, bar in enumerate(bars):
                    height = bar.get_height()
                    if height > 0:
                        y_pos = height / 2
                        x_pos = bar.get_x() + bar.get_width() / 2.
                        font_size = min(12, max(8, int(height / 5)))
                        
                        ax.text(x_pos, y_pos, f'{int(height)}',
                                ha='center', va='center', color='white', 
                                fontsize=font_size, weight='bold',
                                bbox=dict(boxstyle='round,pad=0.3', facecolor='black', alpha=0.8))

                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right', color='white', fontsize=11, weight='bold')
                ax.tick_params(axis='y', colors='white', labelsize=10)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='y', alpha=0.3, color='white')

            elif chart_type == "hbar" and data:
                labels, values = data
                colors = ['#3498DB', '#2CC985', '#F39C12', '#E74C3C', '#9B59B6']
                bars = ax.barh(range(len(labels)), values, color=colors[:len(labels)])
                
                # Add value labels in the middle of horizontal bars
                for i, bar in enumerate(bars):
                    width = bar.get_width()
                    # Position text in the middle of the horizontal bar
                    ax.text(width / 2, bar.get_y() + bar.get_height() / 2.,
                            f'{int(width)}',
                            ha='center', va='center', color='white', fontsize=11, weight='bold',
                            bbox=dict(boxstyle='round,pad=0.3', facecolor='black', alpha=0.7))
                
                ax.set_yticks(range(len(labels)))
                ax.set_yticklabels(labels, color='white', fontsize=11, weight='bold')
                ax.tick_params(axis='x', colors='white', labelsize=10)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='x', alpha=0.3, color='white')

            elif chart_type == "line" and data:
                labels, values = data
                ax.plot(labels, values, marker='o', linewidth=3, markersize=10, color='#3498DB')
                ax.fill_between(range(len(labels)), values, alpha=0.3, color='#3498DB')
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, color='white', fontsize=11, weight='bold')
                ax.tick_params(axis='y', colors='white', labelsize=10)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(True, alpha=0.3, color='white')

            elif chart_type == "hist" and data:
                values, bins = data
                ax.hist(values, bins=bins, color='#3498DB', edgecolor='white', alpha=0.7)
                ax.tick_params(axis='both', colors='white', labelsize=10)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='y', alpha=0.3, color='white')

            elif chart_type == "donut" and data:
                labels, sizes = data
                colors = ['#2CC985', '#E74C3C']
                # Use optimized autopct function from utils
                autopct_func = create_autopct_function(sizes)
                wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct=autopct_func,
                                                  colors=colors, startangle=90)
                # Draw circle for donut
                centre_circle = plt.Circle((0, 0), 0.70, fc='#1e1e1e')
                ax.add_artist(centre_circle)
                for text in texts:
                    text.set_color('white')
                    text.set_fontsize(12)
                    text.set_weight('bold')
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_fontsize(9)
                    autotext.set_weight('bold')
                    autotext.set_fontsize(12)
                    autotext.set_weight('bold')
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)

            elif chart_type == "bar_with_range" and data:
                labels, values, y_range = data
                colors = ['#3498DB', '#2CC985', '#F39C12', '#E74C3C', '#9B59B6']
                bars = ax.bar(range(len(labels)), values, color=colors[:len(labels)])

                # Add value labels in the CENTER of bars
                for i, bar in enumerate(bars):
                    height = bar.get_height()
                    if height > 0:  # Only add label if bar has height
                        # Position text in the center of the bar
                        y_pos = height / 2
                        x_pos = bar.get_x() + bar.get_width() / 2.
                        
                        ax.text(x_pos, y_pos, f'{int(height)}',
                                ha='center', va='center', color='white', 
                                fontsize=12, weight='bold',
                                bbox=dict(boxstyle='round,pad=0.2', facecolor='black', alpha=0.7))

                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right', color='white', fontsize=10, weight='bold')
                ax.set_ylim(0, max(y_range) if y_range else 100)
                ax.tick_params(axis='y', colors='white', labelsize=9)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='y', alpha=0.3, color='white')

            elif chart_type == "grouped_bar" and data:
                branches, total_students, placed_students = data
                x = range(len(branches))
                width = 0.35

                # Create grouped bars
                bars1 = ax.bar([i - width / 2 for i in x], total_students, width, label='Total Students',
                               color='#3498DB')
                bars2 = ax.bar([i + width / 2 for i in x], placed_students, width, label='Placed Students',
                               color='#2CC985')

                # Add value labels in the middle of bars
                for bar in bars1:
                    height = bar.get_height()
                    if height > 0:  # Only show label if bar has height
                        ax.text(bar.get_x() + bar.get_width() / 2., height / 2,
                                f'{int(height)}',
                                ha='center', va='center', color='white', fontsize=10, weight='bold',
                                bbox=dict(boxstyle='round,pad=0.2', facecolor='black', alpha=0.7))

                for bar in bars2:
                    height = bar.get_height()
                    if height > 0:  # Only show label if bar has height
                        ax.text(bar.get_x() + bar.get_width() / 2., height / 2,
                                f'{int(height)}',
                                ha='center', va='center', color='white', fontsize=10, weight='bold',
                                bbox=dict(boxstyle='round,pad=0.2', facecolor='black', alpha=0.7))

                ax.set_xticks(x)
                ax.set_xticklabels(branches, rotation=45, ha='right', color='white', fontsize=11, weight='bold')
                ax.set_ylim(0, 350)
                ax.set_yticks([0, 25, 50, 75, 100, 125, 150, 175, 200, 225, 250, 275, 300, 325, 350])
                ax.tick_params(axis='y', colors='white', labelsize=9)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.legend(loc='upper right', facecolor='#2b2b2b', edgecolor='white', labelcolor='white')
                ax.grid(axis='y', alpha=0.3, color='white')

            # Embed in tkinter
            canvas = FigureCanvasTkAgg(fig, master=frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=5, pady=5)

        except Exception as e:
            error_label = ctk.CTkLabel(frame, text=f"Error: {str(e)[:50]}",
                                       font=("Arial", 10), text_color=COLORS["error"])
            error_label.pack(pady=10)

    def get_students_by_branch(self, students):
        """Get student count by branch"""
        branches = {}
        for student in students:
            # Check both flat and nested structure
            branch = student.get("branch") or student.get("personal_info", {}).get("branch", "Unknown")
            branches[branch] = branches.get(branch, 0) + 1

        if not branches:
            return (["No Data"], [0])

        return (list(branches.keys())[:8], list(branches.values())[:8])

    def get_package_distribution(self, companies):
        """Optimized company package distribution by range (in LPA)"""
        # Define package ranges in LPA
        ranges = {
            "5-10": 0,
            "10-15": 0,
            "15-20": 0,
            "20-25": 0,
            "25-30": 0
        }

        # Optimized single-pass processing
        for company in companies:
            pkg = company.get("package", "0")
            pkg_num = extract_numeric_value(pkg)
            
            # Categorize into ranges efficiently
            if pkg_num < 10:
                ranges["5-10"] += 1
            elif pkg_num < 15:
                ranges["10-15"] += 1
            elif pkg_num < 20:
                ranges["15-20"] += 1
            elif pkg_num < 25:
                ranges["20-25"] += 1
            else:
                ranges["25-30"] += 1

        # Calculate appropriate y-range based on actual values
        max_value = max(ranges.values()) if ranges.values() else 50
        y_max = max(50, int(max_value * 1.1))  # At least 50, or 110% of max value
        y_step = max(5, int(y_max / 10))  # Reasonable step size
        y_range = list(range(0, y_max + y_step, y_step))
        
        return (list(ranges.keys()), list(ranges.values()), y_range)

    def get_placement_stats(self, students, placements):
        """Get placement rate"""
        total_students = len(students)
        placed_students = len(placements)

        if total_students == 0:
            return (["No Students"], [0, 0])

        unplaced = total_students - placed_students
        return (["Placed", "Unplaced"], [placed_students, unplaced])

    def get_top_companies(self, placements):
        """Get top companies by placement count"""
        companies = {}
        for placement in placements:
            company = placement.get("company_name", "Unknown")
            companies[company] = companies.get(company, 0) + 1

        if not companies:
            return (["No Data"], [0])

        sorted_comp = sorted(companies.items(), key=lambda x: x[1], reverse=True)[:8]
        return (list(dict(sorted_comp).keys()), list(dict(sorted_comp).values()))

    def get_records_count(self, students, companies, placements):
        """Get total records count with proper range"""
        labels = ["Students", "Companies", "Placements"]
        values = [len(students), len(companies), len(placements)]
        # Calculate dynamic y_range based on max value
        max_val = max(values) if values else 100
        step = max(25, int(max_val / 10))  # At least 25, or 1/10 of max
        y_range = list(range(0, int(max_val * 1.2) + step, step))
        return (labels, values, y_range)

    def get_package_ranges(self, companies):
        """Get package range distribution"""
        import re
        packages = []

        for company in companies:
            pkg = company.get("package", "0")
            # Extract numeric value
            numbers = re.findall(r'\d+', str(pkg))
            if numbers:
                try:
                    pkg_num = float(numbers[0])
                    if "lpa" in str(pkg).lower():
                        pkg_num = pkg_num * 100000
                    packages.append(pkg_num)
                except:
                    pass

        if not packages:
            return ([0], 5)

        return (packages, 8)

    def get_placements_by_branch(self, placements):
        """Get placements by branch with all branches"""
        branches = {}
        for placement in placements:
            branch = placement.get("student_branch", "Unknown")
            branches[branch] = branches.get(branch, 0) + 1

        if not branches:
            return (["No Data"], [0], [0, 50, 100, 150, 200, 250, 300])

        # Sort by branch name for consistency
        sorted_branch = sorted(branches.items(), key=lambda x: x[0])
        return (list(dict(sorted_branch).keys()), list(dict(sorted_branch).values()), [0, 50, 100, 150, 200, 250, 300])

    def get_placements_by_hr(self, placements):
        """Get placements by HR"""
        hrs = {}
        for placement in placements:
            hr = placement.get("hr_name", "Unknown")
            hrs[hr] = hrs.get(hr, 0) + 1

        if not hrs:
            return (["No Data"], [0])

        sorted_hr = sorted(hrs.items(), key=lambda x: x[1], reverse=True)[:8]
        return (list(dict(sorted_hr).keys()), list(dict(sorted_hr).values()))

    def get_placement_rate_by_branch(self, students, placements):
        """Get placement rate percentage by branch"""
        # Count students by branch
        student_branches = {}
        for student in students:
            branch = student.get("branch", "Unknown")
            student_branches[branch] = student_branches.get(branch, 0) + 1

        # Count placed students by branch
        placed_branches = {}
        for placement in placements:
            branch = placement.get("student_branch", "Unknown")
            placed_branches[branch] = placed_branches.get(branch, 0) + 1

        # Calculate placement rates
        branches = sorted(set(list(student_branches.keys()) + list(placed_branches.keys())))
        rates = []
        for branch in branches:
            total = student_branches.get(branch, 0)
            placed = placed_branches.get(branch, 0)
            rate = (placed / total * 100) if total > 0 else 0
            rates.append(rate)

        return (branches, rates)

    def get_package_wise_placements(self, placements):
        """Get number of placements by package"""
        packages = {}
        for placement in placements:
            pkg = placement.get("package", "Unknown")
            packages[pkg] = packages.get(pkg, 0) + 1

        if not packages:
            return (["No Data"], [0])

        sorted_pkg = sorted(packages.items(), key=lambda x: x[1], reverse=True)[:8]
        return (list(dict(sorted_pkg).keys()), list(dict(sorted_pkg).values()))

    def get_company_wise_placements(self, placements):
        """Get top companies by placement count"""
        companies = {}
        for placement in placements:
            company = placement.get("company_name", "Unknown")
            companies[company] = companies.get(company, 0) + 1

        if not companies:
            return (["No Data"], [0])

        sorted_comp = sorted(companies.items(), key=lambda x: x[1], reverse=True)[:10]
        return (list(dict(sorted_comp).keys()), list(dict(sorted_comp).values()))

    def get_avg_package_by_branch(self, placements):
        """Get average package by branch"""
        import re
        branch_packages = {}

        for placement in placements:
            branch = placement.get("student_branch", "Unknown")
            pkg = placement.get("package", "0")

            numbers = re.findall(r'\d+', str(pkg))
            if numbers:
                try:
                    pkg_num = float(numbers[0])
                    if "lpa" in str(pkg).lower():
                        pkg_num = pkg_num

                    if branch not in branch_packages:
                        branch_packages[branch] = []
                    branch_packages[branch].append(pkg_num)
                except:
                    pass

        branches = sorted(branch_packages.keys())
        avg_packages = []
        for branch in branches:
            avg = sum(branch_packages[branch]) / len(branch_packages[branch])
            avg_packages.append(round(avg, 2))

        return (branches, avg_packages)

    def get_students_vs_placed_by_branch(self, students, placements):
        """Get students vs placed students by branch"""
        # Count students by branch
        student_branches = {}
        for student in students:
            # Check both flat and nested structure
            branch = student.get("branch") or student.get("personal_info", {}).get("branch", "Unknown")
            student_branches[branch] = student_branches.get(branch, 0) + 1

        # Count placed students by branch
        placed_branches = {}
        for placement in placements:
            branch = placement.get("student_branch", "Unknown")
            placed_branches[branch] = placed_branches.get(branch, 0) + 1

        # Get all branches sorted
        branches = sorted(set(list(student_branches.keys()) + list(placed_branches.keys())))

        # Prepare data for grouped bar chart
        total_students = []
        placed_students = []

        for branch in branches:
            total_students.append(student_branches.get(branch, 0))
            placed_students.append(placed_branches.get(branch, 0))

        return (branches, total_students, placed_students)

    def export_home_charts_to_excel(self):
        """Export home dashboard chart data to Excel with charts"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.chart.series import DataPoint
            from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
            from datetime import datetime
            from tkinter import filedialog

            # Get data from databases
            students = []
            companies = []
            placements = []

            if self.student_manager.collection is not None:
                students = list(self.student_manager.collection.find())
            if self.company_manager.collection is not None:
                companies = list(self.company_manager.collection.find())
            if self.placed_student_manager.collection is not None:
                placements = list(self.placed_student_manager.collection.find())

            if not students and not companies and not placements:
                messagebox.showwarning("No Data", "No data available to export.")
                return

            # Open Save As dialog
            default_filename = f"home_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Dashboard Data As"
            )

            if not filename:  # User cancelled
                return

            # Create workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Header styles
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Chart colors
            chart_colors = ['3498DB', '2CC985', 'F39C12', 'E74C3C', '9B59B6', '1ABC9C', 'E67E22', '34495E']

            # ========== Sheet 1: Students by Branch (Pie Chart) ==========
            ws1 = wb.create_sheet("Students by Branch")
            ws1.append(["Branch", "Count"])

            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            branches_data = self.get_students_by_branch(students)
            if branches_data:
                labels, values = branches_data
                for label, value in zip(labels, values):
                    ws1.append([label, value])

            # Style data cells
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 15

            # Create Pie Chart for Students by Branch
            if branches_data and len(branches_data[0]) > 0:
                pie_chart = PieChart()
                pie_chart.title = "Students by Branch"
                # Data starts from row 2 (skip header) to avoid header label in chart
                data = Reference(ws1, min_col=2, min_row=2, max_row=len(branches_data[0]) + 1)
                cats = Reference(ws1, min_col=1, min_row=2, max_row=len(branches_data[0]) + 1)
                pie_chart.add_data(data, titles_from_data=False)
                pie_chart.set_categories(cats)
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.dataLabels.showCatName = True
                pie_chart.dataLabels.showVal = False
                pie_chart.dataLabels.showSerName = False
                pie_chart.width = 15
                pie_chart.height = 10
                ws1.add_chart(pie_chart, "D2")

            # ========== Sheet 2: Package Distribution (Bar Chart) ==========
            ws2 = wb.create_sheet("Package Distribution")
            ws2.append(["Package Range (LPA)", "Count"])

            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            pkg_data = self.get_package_distribution(companies)
            if pkg_data:
                labels, values, _ = pkg_data
                for label, value in zip(labels, values):
                    ws2.append([label, value])

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 15

            # Create Bar Chart for Package Distribution
            if pkg_data and len(pkg_data[0]) > 0:
                bar_chart = BarChart()
                bar_chart.title = "Company Package Ranges (LPA)"
                bar_chart.type = "col"
                bar_chart.style = 10
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(pkg_data[0]) + 1)
                cats = Reference(ws2, min_col=1, min_row=2, max_row=len(pkg_data[0]) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(cats)
                bar_chart.dataLabels = DataLabelList()
                bar_chart.dataLabels.showVal = True
                bar_chart.width = 15
                bar_chart.height = 10
                ws2.add_chart(bar_chart, "D2")

            # ========== Sheet 3: Total Records (Bar Chart) ==========
            ws3 = wb.create_sheet("Total Records")
            ws3.append(["Category", "Count"])

            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            records_data = self.get_records_count(students, companies, placements)
            if records_data:
                labels, values, _ = records_data
                for label, value in zip(labels, values):
                    ws3.append([label, value])

            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws3.column_dimensions['A'].width = 20
            ws3.column_dimensions['B'].width = 15

            # Create Bar Chart for Total Records
            if records_data and len(records_data[0]) > 0:
                bar_chart2 = BarChart()
                bar_chart2.title = "Total Records Count"
                bar_chart2.type = "col"
                bar_chart2.style = 10
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(records_data[0]) + 1)
                cats = Reference(ws3, min_col=1, min_row=2, max_row=len(records_data[0]) + 1)
                bar_chart2.add_data(data, titles_from_data=True)
                bar_chart2.set_categories(cats)
                bar_chart2.dataLabels = DataLabelList()
                bar_chart2.dataLabels.showVal = True
                bar_chart2.width = 15
                bar_chart2.height = 10
                ws3.add_chart(bar_chart2, "D2")

            # ========== Sheet 4: Students vs Placed (Grouped Bar Chart) ==========
            ws4 = wb.create_sheet("Students vs Placed")
            ws4.append(["Branch", "Total Students", "Placed Students", "Placement Rate (%)"])

            for cell in ws4[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            vs_data = self.get_students_vs_placed_by_branch(students, placements)
            if vs_data:
                branches, total_students, placed_students = vs_data
                for branch, total, placed in zip(branches, total_students, placed_students):
                    rate = (placed / total * 100) if total > 0 else 0
                    ws4.append([branch, total, placed, f"{rate:.1f}%"])

            for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws4.column_dimensions['A'].width = 20
            ws4.column_dimensions['B'].width = 18
            ws4.column_dimensions['C'].width = 18
            ws4.column_dimensions['D'].width = 20

            # Create Grouped Bar Chart for Students vs Placed
            if vs_data and len(vs_data[0]) > 0:
                grouped_bar = BarChart()
                grouped_bar.title = "Students vs Placed by Branch"
                grouped_bar.type = "col"
                grouped_bar.style = 10
                grouped_bar.grouping = "clustered"
                
                # Add both data series (Total Students and Placed Students)
                data = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=len(vs_data[0]) + 1)
                cats = Reference(ws4, min_col=1, min_row=2, max_row=len(vs_data[0]) + 1)
                grouped_bar.add_data(data, titles_from_data=True)
                grouped_bar.set_categories(cats)
                grouped_bar.dataLabels = DataLabelList()
                grouped_bar.dataLabels.showVal = True
                grouped_bar.width = 18
                grouped_bar.height = 10
                ws4.add_chart(grouped_bar, "F2")

            # Save file
            wb.save(filename)
            
            # Count charts added
            chart_count = 0
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                chart_count += len(ws._charts)
            
            messagebox.showinfo("Success", 
                f"Dashboard exported successfully to:\n{filename}\n\n"
                f"📊 Charts added: {chart_count}\n"
                f"📋 Sheets: {len(wb.sheetnames)}")

        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed. Install it using: pip install openpyxl")
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("Error", f"Failed to export to Excel: {e}\n\nDetails:\n{error_details[:500]}")

    def toggle_maximize(self):
        if self.is_maximized:
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            width = int(screen_width * 0.8)
            height = int(screen_height * 0.8)
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            self.normal_geometry = f"{width}x{height}+{x}+{y}"
            self.root.geometry(self.normal_geometry)
            self.maximize_button.configure(text="□")
            self.is_maximized = False
        else:
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            self.root.geometry(f"{screen_width}x{screen_height}+0+0")
            self.maximize_button.configure(text="❐")
            self.is_maximized = True

    def minimize_window(self):
        self.root.iconify()

    def open_database_settings(self):
        """Open MongoDB Atlas in browser"""
        import webbrowser

        # MongoDB Atlas URL
        mongodb_url = "https://cloud.mongodb.com"

        try:
            # Open the URL in default browser
            webbrowser.open(mongodb_url)
            messagebox.showinfo("Database Server",
                                "Opening MongoDB Atlas in your browser...\n\n"
                                "URL: " + mongodb_url)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open browser: {e}\n\n"
                                          "Please visit: https://cloud.mongodb.com")

    def close_window(self):
        if messagebox.askyesno("Exit", "Are you sure you want to exit the application?"):
            self.root.destroy()


def start_app():
    """Start the application"""
    root = ctk.CTk()
    app = TP_Manager(root)
    root.mainloop()


if __name__ == "__main__":
    start_app()