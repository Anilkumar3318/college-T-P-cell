import customtkinter as ctk
from tkinter import messagebox
from database_config import get_placed_student_collection, get_offer_letters_collection
from pdf_manager import pdf_manager
import datetime
import os
from utils import (
    validate_email, validate_phone, uppercase_entry_handler,
    build_search_query, batch_clear_widgets, safe_int_convert, 
    safe_float_convert, format_filter_info, create_autopct_function,
    extract_numeric_value, COLORS, CHART_COLORS, invalidate_cache,
    performance_monitor
)


class PlacedStudentManager:
    def __init__(self):
        self.collection = self.connect_db()
        self.offer_letters_collection = self.connect_offer_letters_db()

    def connect_db(self):
        return get_placed_student_collection()

    def connect_offer_letters_db(self):
        return get_offer_letters_collection()

    def uppercase_entry(self, entry_widget):
        return uppercase_entry_handler(entry_widget)

    def show_add_interface(self, parent):
        """Show add placed student interface"""
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_add_placed_student_tab(parent)

    def show_view_interface(self, parent):
        """Show view placed students interface"""
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_view_placed_students_tab(parent)

    def show_edit_interface(self, parent):
        """Show edit placed student interface"""
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_edit_placed_student_tab(parent)

    def show_delete_interface(self, parent):
        """Show delete placed student interface"""
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_delete_placed_student_tab(parent)

    def show_chart_interface(self, parent):
        """Show placed student analytics charts"""
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_placed_charts_tab(parent)

    def setup_placed_charts_tab(self, parent):
        """Setup placed student analytics charts"""
        # Header with title and export button
        header_frame = ctk.CTkFrame(parent, fg_color=COLORS["content_frame"])
        header_frame.pack(fill='x', padx=5, pady=3)

        title = ctk.CTkLabel(header_frame, text="📊 PLACEMENT ANALYTICS",
                             font=("Arial", 18, "bold"), text_color=COLORS["info"])
        title.pack(side='left', padx=10, pady=5)

        export_btn = ctk.CTkButton(header_frame, text="📊 EXPORT TO EXCEL",
                                   command=self.export_placed_charts_to_excel,
                                   height=32, width=150, font=("Arial", 11, "bold"),
                                   fg_color=COLORS["success"])
        export_btn.pack(side='right', padx=10, pady=5)

        # Input Frame for filters - compact single row
        input_frame = ctk.CTkFrame(parent, fg_color=COLORS["section_frame"], corner_radius=10)
        input_frame.pack(fill='x', padx=5, pady=3)

        ctk.CTkLabel(input_frame, text="Batch Year:", font=("Arial", 11, "bold"),
                     text_color="white").pack(side='left', padx=3, pady=5)

        if not hasattr(self, 'placed_batch_year_var'):
            self.placed_batch_year_var = ctk.StringVar(value="")

        batch_entry = ctk.CTkEntry(input_frame, textvariable=self.placed_batch_year_var, width=80, height=32,
                                   placeholder_text="2022", font=("Arial", 11))
        batch_entry.pack(side='left', padx=3, pady=5)

        ctk.CTkLabel(input_frame, text="Branch:", font=("Arial", 11, "bold"),
                     text_color="white").pack(side='left', padx=3, pady=5)

        if not hasattr(self, 'placed_branch_filter_var'):
            self.placed_branch_filter_var = ctk.StringVar(value="All Branches")

        branch_options = ["All Branches"]
        if self.collection is not None:
            try:
                all_placements = list(self.collection.find())
                branches = set()
                for placement in all_placements:
                    branch = placement.get("student_branch", "")
                    if branch:
                        branches.add(branch)
                branch_options.extend(sorted(branches))
            except:
                pass

        branch_dropdown = ctk.CTkOptionMenu(input_frame, variable=self.placed_branch_filter_var, 
                                            values=branch_options, width=130, height=32,
                                            font=("Arial", 11))
        branch_dropdown.pack(side='left', padx=3, pady=5)

        search_btn = ctk.CTkButton(input_frame, text="🔍 SEARCH",
                                   command=lambda: self.refresh_placed_charts(parent),
                                   height=32, width=100, font=("Arial", 11, "bold"),
                                   fg_color=COLORS["info"])
        search_btn.pack(side='left', padx=5, pady=5)

        # Create scrollable frame for charts
        charts_container = ctk.CTkFrame(parent, fg_color=COLORS["content_frame"])
        charts_container.pack(fill='both', expand=True, padx=5, pady=3)

        # Loading indicator
        loading_label = ctk.CTkLabel(charts_container, text="⏳ Loading analytics...",
                                     font=("Arial", 14), text_color=COLORS["info"])
        loading_label.pack(pady=50)
        charts_container.update()

        try:
            # Get all placements with pagination and projection
            all_placements = []
            if self.collection is not None:
                try:
                    projection = {
                        "student_name": 1, "student_branch": 1, "company_name": 1,
                        "package": 1, "hr_name": 1, "placement_date": 1
                    }
                    all_placements = list(self.collection.find({}, projection)
                                        .sort([("_id", -1)])
                                        .limit(500))  # Limit for performance
                except Exception as e:
                    print(f"Error fetching placements: {e}")
                    all_placements = []

            # Filter placements
            placements = all_placements
            selected_year = self.placed_batch_year_var.get().strip()
            selected_branch = self.placed_branch_filter_var.get() if hasattr(self, 'placed_branch_filter_var') else "All Branches"
            filter_parts = []
            status_color = COLORS["info"]

            # Apply batch year filter
            if selected_year:
                try:
                    year_int = int(selected_year)
                    year_str = str(year_int)
                    
                    # Get students from that batch
                    from database_config import get_student_collection
                    student_collection = get_student_collection()
                    if student_collection is not None:
                        # Find students with matching admission year (check both int and string)
                        batch_student_names = set()
                        all_students = list(student_collection.find())
                        for s in all_students:
                            admission_year = s.get("admission_year") or s.get("personal_info", {}).get("admission_year")
                            if admission_year is not None:
                                if str(admission_year) == year_str or admission_year == year_int:
                                    name = s.get("name") or s.get("personal_info", {}).get("name", "")
                                    if name:
                                        batch_student_names.add(name.upper())
                        
                        placements = [p for p in placements if p.get("student_name", "").upper() in batch_student_names]
                        filter_parts.append(f"Batch {year_int}")
                except ValueError:
                    filter_parts.append("Invalid year")
                    status_color = COLORS["error"]

            # Apply branch filter
            if selected_branch and selected_branch != "All Branches":
                placements = [p for p in placements if p.get("student_branch", "").upper() == selected_branch.upper()]
                filter_parts.append(f"Branch: {selected_branch}")

            # Build filter status message
            if filter_parts:
                filter_status = f"{' | '.join(filter_parts)} - {len(placements)} placements"
                if len(placements) > 0:
                    status_color = COLORS["success"]
                else:
                    status_color = COLORS["warning"]
            else:
                filter_status = f"All Data - {len(all_placements)} placements"
                status_color = COLORS["info"]

            # Remove loading label
            loading_label.destroy()

            # Show filter status
            status_label = ctk.CTkLabel(charts_container, text=f"📊 {filter_status}",
                                        font=("Arial", 13, "bold"), text_color=status_color)
            status_label.pack(pady=10)

            if not placements:
                error_label = ctk.CTkLabel(charts_container, text="❌ No placement data to display.",
                                           font=("Arial", 12), text_color=COLORS["error"])
                error_label.pack(pady=50)
                return

            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.figure import Figure

            # Row 1: 2 charts
            row1_frame = ctk.CTkFrame(charts_container, fg_color=COLORS["content_frame"])
            row1_frame.pack(fill='both', expand=True, padx=5, pady=5)

            # Chart 1: Placements by Branch
            self.create_placed_chart(row1_frame, "Placements by Branch",
                                     self.get_placements_by_branch(placements), "pie", row1_frame)

            # Chart 2: Placements by Package
            self.create_placed_chart(row1_frame, "Placements by Package",
                                     self.get_placements_by_package(placements), "bar", row1_frame)

            # Row 2: 2 charts
            row2_frame = ctk.CTkFrame(charts_container, fg_color=COLORS["content_frame"])
            row2_frame.pack(fill='both', expand=True, padx=5, pady=5)

            # Chart 3: Top Companies
            self.create_placed_chart(row2_frame, "Top 10 Companies by Placements",
                                     self.get_top_companies(placements), "bar", row2_frame)

            # Chart 4: Placements by HR
            self.create_placed_chart(row2_frame, "Placements by HR",
                                     self.get_placements_by_hr(placements), "pie", row2_frame)

            # Store chart data for export
            self.current_chart_data = {
                'placements': placements,
                'filter_status': filter_status
            }

        except Exception as e:
            loading_label.destroy()
            error_label = ctk.CTkLabel(charts_container, text=f"Error loading analytics: {e}",
                                       font=("Arial", 14), text_color=COLORS["error"])
            error_label.pack(pady=50)

    def refresh_placed_charts(self, parent):
        """Refresh charts with selected filter"""
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_placed_charts_tab(parent)

    def create_placed_chart(self, parent, title, data, chart_type, row_frame):
        """Create a placement chart"""
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.figure import Figure

        frame = ctk.CTkFrame(row_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        frame.pack(side='left', fill='both', expand=True, padx=5, pady=5)

        title_label = ctk.CTkLabel(frame, text=title, font=("Arial", 14, "bold"),
                                   text_color=COLORS["info"])
        title_label.pack(pady=10)

        try:
            fig = Figure(figsize=(6.5, 5), dpi=100, facecolor='#2b2b2b', edgecolor='none')
            ax = fig.add_subplot(111)
            ax.set_facecolor('#1e1e1e')
            fig.subplots_adjust(left=0.12, right=0.95, top=0.88, bottom=0.18)

            if chart_type == "pie" and data:
                labels, sizes = data
                colors = ['#3498DB', '#2CC985', '#F39C12', '#E74C3C', '#9B59B6', '#1ABC9C', '#E67E22', '#34495E']
                # Custom autopct function to show both value and percentage
                def make_autopct(values):
                    def autopct(pct):
                        total = sum(values)
                        val = int(round(pct * total / 100.0))
                        return f'{val}\n({pct:.1f}%)'
                    return autopct
                ax.pie(sizes, labels=labels, autopct=make_autopct(sizes), colors=colors[:len(labels)],
                       textprops={'color': 'white', 'fontsize': 9, 'weight': 'bold'})
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)

            elif chart_type == "bar" and data:
                labels, values = data
                colors = ['#3498DB', '#2CC985', '#F39C12', '#E74C3C', '#9B59B6']
                bars = ax.bar(range(len(labels)), values, color=colors[:len(labels)])

                # Add value labels in the middle of bars for better visibility
                for i, bar in enumerate(bars):
                    height = bar.get_height()
                    if height > 0:  # Only add label if bar has height
                        # Position text in the middle of the bar
                        y_pos = height / 2
                        x_pos = bar.get_x() + bar.get_width() / 2.
                        
                        # Adjust font size based on bar height for better fit
                        font_size = min(12, max(8, int(height / 5)))
                        
                        ax.text(x_pos, y_pos, f'{int(height)}',
                                ha='center', va='center', color='white', 
                                fontsize=font_size, weight='bold',
                                bbox=dict(boxstyle='round,pad=0.3', facecolor='black', alpha=0.8))

                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right', color='white', fontsize=11, weight='bold')
                ax.tick_params(axis='y', colors='white', labelsize=10)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='y', alpha=0.3, color='white')

            canvas = FigureCanvasTkAgg(fig, master=frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=5, pady=5)

        except Exception as e:
            error_label = ctk.CTkLabel(frame, text=f"Error: {str(e)[:50]}",
                                       font=("Arial", 10), text_color=COLORS["error"])
            error_label.pack(pady=10)

    def get_placements_by_branch(self, placements):
        """Get placements by branch"""
        branches = {}
        for placement in placements:
            branch = placement.get("student_branch", "Unknown")
            branches[branch] = branches.get(branch, 0) + 1

        if not branches:
            return (["No Data"], [0])

        return (list(branches.keys()), list(branches.values()))

    def get_placements_by_package(self, placements):
        """Get placements by package"""
        packages = {}
        for placement in placements:
            pkg = placement.get("package", "Unknown")
            packages[pkg] = packages.get(pkg, 0) + 1

        if not packages:
            return (["No Data"], [0])

        sorted_pkg = sorted(packages.items(), key=lambda x: x[1], reverse=True)[:8]
        return (list(dict(sorted_pkg).keys()), list(dict(sorted_pkg).values()))

    def get_top_companies(self, placements):
        """Get top companies by placement count"""
        companies = {}
        for placement in placements:
            company = placement.get("company_name", "Unknown")
            companies[company] = companies.get(company, 0) + 1

        if not companies:
            return (["No Data"], [0])

        sorted_comp = sorted(companies.items(), key=lambda x: x[1], reverse=True)[:10]
        return (list(dict(sorted_comp).keys()), list(dict(sorted_comp).values()))

    def get_placements_by_hr(self, placements):
        """Get placements by HR"""
        hrs = {}
        for placement in placements:
            hr = placement.get("hr_name", "Unknown")
            hrs[hr] = hrs.get(hr, 0) + 1

        if not hrs:
            return (["No Data"], [0])

        sorted_hr = sorted(hrs.items(), key=lambda x: x[1], reverse=True)[:8]
        return (list(dict(sorted_hr).keys()), list(dict(sorted_hr).values()))

    def setup_add_placed_student_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="➕ ADD PLACED STUDENT",
                                   font=("Arial", 18, "bold"), text_color=COLORS["success"])
        title_label.pack(pady=5)

        # Placed Student Information Section - Compact 2-column layout
        placed_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        placed_frame.pack(fill='x', padx=10, pady=5)

        ctk.CTkLabel(placed_frame, text="🎓 PLACED STUDENT INFORMATION",
                     font=("Arial", 14, "bold"), text_color=COLORS["info"]).pack(pady=8)

        # Placed student info fields - 2 column grid
        placed_grid = ctk.CTkFrame(placed_frame, fg_color=COLORS["section_frame"])
        placed_grid.pack(fill='x', padx=15, pady=8)

        self.placed_company_name_var = ctk.StringVar()
        self.placed_email_var = ctk.StringVar()
        self.placed_contact_info_var = ctk.StringVar()
        self.placed_hr_name_var = ctk.StringVar()
        self.placed_package_var = ctk.StringVar()
        self.placed_address_var = ctk.StringVar()
        self.student_name_var = ctk.StringVar()
        self.student_branch_var = ctk.StringVar()
        self.year_of_placement_var = ctk.StringVar()
        self.position_var = ctk.StringVar()
        self.batch_var = ctk.StringVar()
        self.offer_letter_path_var = ctk.StringVar()
        self.offer_letter_key_var = ctk.StringVar()
        self.company_levels_var = ctk.StringVar()
        self.skills_required_var = ctk.StringVar()
        self.important_suggestions_var = ctk.StringVar()

        # Row 1: Student Name, Branch
        ctk.CTkLabel(placed_grid, text="Student Name*:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=4)
        name_entry = ctk.CTkEntry(placed_grid, textvariable=self.student_name_var, width=200, height=32)
        name_entry.grid(row=0, column=1, sticky='w', padx=5, pady=4)
        self.student_name_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(placed_grid, text="Branch*:", font=("Arial", 11, "bold")).grid(row=0, column=2, sticky='w', padx=5, pady=4)
        branch_entry = ctk.CTkEntry(placed_grid, textvariable=self.student_branch_var, width=200, height=32)
        branch_entry.grid(row=0, column=3, sticky='w', padx=5, pady=4)
        self.student_branch_var.trace_add('write', self.uppercase_entry(branch_entry))

        # Row 2: Batch, Company
        ctk.CTkLabel(placed_grid, text="Batch*:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky='w', padx=5, pady=4)
        batch_entry = ctk.CTkEntry(placed_grid, textvariable=self.batch_var, width=200, height=32)
        batch_entry.grid(row=1, column=1, sticky='w', padx=5, pady=4)
        self.batch_var.trace_add('write', self.uppercase_entry(batch_entry))

        ctk.CTkLabel(placed_grid, text="Company*:", font=("Arial", 11, "bold")).grid(row=1, column=2, sticky='w', padx=5, pady=4)
        company_entry = ctk.CTkEntry(placed_grid, textvariable=self.placed_company_name_var, width=200, height=32)
        company_entry.grid(row=1, column=3, sticky='w', padx=5, pady=4)
        self.placed_company_name_var.trace_add('write', self.uppercase_entry(company_entry))

        # Row 3: Position, Year of Placement
        ctk.CTkLabel(placed_grid, text="Position*:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', padx=5, pady=4)
        position_entry = ctk.CTkEntry(placed_grid, textvariable=self.position_var, width=200, height=32)
        position_entry.grid(row=2, column=1, sticky='w', padx=5, pady=4)
        self.position_var.trace_add('write', self.uppercase_entry(position_entry))

        ctk.CTkLabel(placed_grid, text="Year of Placement*:", font=("Arial", 11, "bold")).grid(row=2, column=2, sticky='w', padx=5, pady=4)
        year_entry = ctk.CTkEntry(placed_grid, textvariable=self.year_of_placement_var, width=200, height=32)
        year_entry.grid(row=2, column=3, sticky='w', padx=5, pady=4)

        # Row 4: Package, Email
        ctk.CTkLabel(placed_grid, text="Package*:", font=("Arial", 11, "bold")).grid(row=3, column=0, sticky='w', padx=5, pady=4)
        package_entry = ctk.CTkEntry(placed_grid, textvariable=self.placed_package_var, width=200, height=32)
        package_entry.grid(row=3, column=1, sticky='w', padx=5, pady=4)
        self.placed_package_var.trace_add('write', self.uppercase_entry(package_entry))

        ctk.CTkLabel(placed_grid, text="Email*:", font=("Arial", 11, "bold")).grid(row=3, column=2, sticky='w', padx=5, pady=4)
        ctk.CTkEntry(placed_grid, textvariable=self.placed_email_var, width=200, height=32).grid(row=3, column=3, sticky='w', padx=5, pady=4)

        # Row 5: Contact, HR Name
        ctk.CTkLabel(placed_grid, text="Contact*:", font=("Arial", 11, "bold")).grid(row=4, column=0, sticky='w', padx=5, pady=4)
        ctk.CTkEntry(placed_grid, textvariable=self.placed_contact_info_var, width=200, height=32).grid(row=4, column=1, sticky='w', padx=5, pady=4)

        ctk.CTkLabel(placed_grid, text="HR Name*:", font=("Arial", 11, "bold")).grid(row=4, column=2, sticky='w', padx=5, pady=4)
        hr_entry = ctk.CTkEntry(placed_grid, textvariable=self.placed_hr_name_var, width=200, height=32)
        hr_entry.grid(row=4, column=3, sticky='w', padx=5, pady=4)
        self.placed_hr_name_var.trace_add('write', self.uppercase_entry(hr_entry))

        # Row 6: Address (full width)
        ctk.CTkLabel(placed_grid, text="Address:", font=("Arial", 11, "bold")).grid(row=5, column=0, sticky='w', padx=5, pady=4)
        address_entry = ctk.CTkEntry(placed_grid, textvariable=self.placed_address_var, width=500, height=32)
        address_entry.grid(row=5, column=1, columnspan=3, sticky='w', padx=5, pady=4)
        self.placed_address_var.trace_add('write', self.uppercase_entry(address_entry))

        # Initialize new variables
        self.offer_letter_path_var = ctk.StringVar()
        self.offer_letter_key_var = ctk.StringVar()
        self.placement_suggestion_var = ctk.StringVar()
        self.company_levels_var = ctk.StringVar()
        self.skills_required_var = ctk.StringVar()
        self.important_suggestions_var = ctk.StringVar()

        # Upload Offer Letter Section
        upload_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        upload_frame.pack(fill='x', padx=10, pady=5)

        ctk.CTkLabel(upload_frame, text="📄 UPLOAD OFFER LETTER",
                     font=("Arial", 14, "bold"), text_color=COLORS["info"]).pack(pady=8)

        upload_grid = ctk.CTkFrame(upload_frame, fg_color=COLORS["section_frame"])
        upload_grid.pack(fill='x', padx=15, pady=8)

        ctk.CTkLabel(upload_grid, text="Offer Letter (PDF, max 100KB):", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=4)
        
        upload_row = ctk.CTkFrame(upload_grid, fg_color=COLORS["section_frame"])
        upload_row.grid(row=0, column=1, columnspan=3, sticky='w', padx=5, pady=4)
        
        self.offer_letter_entry = ctk.CTkEntry(upload_row, textvariable=self.offer_letter_path_var, width=300, height=32, 
                                               placeholder_text="No file selected", state="readonly")
        self.offer_letter_entry.pack(side='left', padx=2)
        
        ctk.CTkButton(upload_row, text="📁 Browse", command=self.browse_offer_letter,
                      fg_color=COLORS["info"], font=("Arial", 10, "bold"), height=32, width=80).pack(side='left', padx=2)
        
        ctk.CTkButton(upload_row, text="🗑️ Clear", command=self.clear_offer_letter,
                      fg_color=COLORS["error"], font=("Arial", 10, "bold"), height=32, width=60).pack(side='left', padx=2)

        # Company Placement Suggestions Section
        suggestions_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        suggestions_frame.pack(fill='x', padx=10, pady=5)

        ctk.CTkLabel(suggestions_frame, text="💡 SUGGESTIONS TO PLACE IN THIS COMPANY",
                     font=("Arial", 14, "bold"), text_color=COLORS["warning"]).pack(pady=8)

        suggestions_grid = ctk.CTkFrame(suggestions_frame, fg_color=COLORS["section_frame"])
        suggestions_grid.pack(fill='x', padx=15, pady=8)

        # Row 1: Placement Suggestion (full width)
        ctk.CTkLabel(suggestions_grid, text="Placement Suggestion:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=4)
        suggestion_entry = ctk.CTkEntry(suggestions_grid, textvariable=self.placement_suggestion_var, width=600, height=32,
                                        placeholder_text="suggest how to place in this compony")
        suggestion_entry.grid(row=0, column=1, columnspan=3, sticky='w', padx=5, pady=4)
        self.placement_suggestion_var.trace_add('write', self.uppercase_entry(suggestion_entry))

        # Row 2: Company Levels, Skills Required
        ctk.CTkLabel(suggestions_grid, text="How many levels:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky='w', padx=5, pady=4)
        levels_entry = ctk.CTkEntry(suggestions_grid, textvariable=self.company_levels_var, width=200, height=32,
                                    placeholder_text="e.g., 3 levels, 5 rounds")
        levels_entry.grid(row=1, column=1, sticky='w', padx=5, pady=4)
        self.company_levels_var.trace_add('write', self.uppercase_entry(levels_entry))

        ctk.CTkLabel(suggestions_grid, text="Skills Required:", font=("Arial", 11, "bold")).grid(row=1, column=2, sticky='w', padx=5, pady=4)
        skills_entry = ctk.CTkEntry(suggestions_grid, textvariable=self.skills_required_var, width=300, height=32,
                                    placeholder_text="e.g., Java, Python, Communication")
        skills_entry.grid(row=1, column=3, sticky='w', padx=5, pady=4)
        self.skills_required_var.trace_add('write', self.uppercase_entry(skills_entry))

        # Row 3: Important Suggestions (full width)
        ctk.CTkLabel(suggestions_grid, text="Important Suggestions:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='nw', padx=5, pady=4)
        suggestions_textbox = ctk.CTkTextbox(suggestions_grid, width=600, height=80, font=("Arial", 11))
        suggestions_textbox.grid(row=2, column=1, columnspan=3, sticky='w', padx=5, pady=4)
        suggestions_textbox.bind("<KeyRelease>", lambda e: self.important_suggestions_var.set(suggestions_textbox.get("1.0", "end-1c")))

        # Submit button - compact
        submit_btn = ctk.CTkButton(tab, text="➕ ADD PLACED STUDENT", command=self.add_placed_student,
                                   fg_color=COLORS["success"], font=("Arial", 14, "bold"), height=40, width=180)
        submit_btn.pack(pady=15)

    def browse_offer_letter(self):
        """Browse and select offer letter PDF file"""
        from tkinter import filedialog
        import os
        
        file_path = filedialog.askopenfilename(
            title="Select Offer Letter PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialdir=os.path.expanduser("~")
        )
        
        if file_path:
            # Check file size (max 100KB = 102400 bytes)
            file_size = os.path.getsize(file_path)
            if file_size > 150000:
                messagebox.showerror("File Too Large", 
                    f"File size is {file_size/1500:.1f} KB. Maximum allowed is 150 KB.\n"
                    "Please compress the PDF or select a smaller file.")
                return
            
            # Generate unique key for this PDF that will link to the placement record
            import uuid
            import datetime
            unique_key = f"offer_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4())[:8]}"
            
            # Store the file path and key for later use
            self.offer_letter_path_var.set(file_path)
            self.offer_letter_key_var.set(unique_key)
            
            # Update entry to show filename
            filename = os.path.basename(file_path)
            self.offer_letter_entry.configure(placeholder_text=f"✓ {filename} ({file_size/1024:.1f} KB)")
            
            messagebox.showinfo("File Selected", 
                f"File: {filename}\n"
                f"Size: {file_size/1500:.1f} KB\n"
                f"Unique Key: {unique_key}\n\n"
                f"Note: PDF will be stored in separate database when you submit the form.")

    def store_offer_letter_in_db(self, pdf_key, file_path, student_name, company_name):
        """Store offer letter PDF in separate database with linking key"""
        try:
            if not self.offer_letters_collection:
                return False
                
            # Read PDF file as binary data
            with open(file_path, 'rb') as pdf_file:
                pdf_data = pdf_file.read()
            
            # Create offer letter document with linking information
            offer_letter_doc = {
                "pdf_key": pdf_key,  # Unique key to link with placement record
                "student_name": student_name,
                "company_name": company_name,
                "filename": os.path.basename(file_path),
                "file_size": len(pdf_data),
                "pdf_data": pdf_data,  # Binary PDF data
                "upload_date": datetime.datetime.now(),
                "file_type": "application/pdf"
            }
            
            # Store in offer letters database
            result = self.offer_letters_collection.insert_one(offer_letter_doc)
            
            if result.inserted_id:
                return True
            else:
                return False
                
        except Exception as e:
            messagebox.showerror("Storage Error", f"Failed to store offer letter: {e}")
            return False

    def clear_offer_letter(self):
        """Clear selected offer letter"""
        self.offer_letter_path_var.set("")
        self.offer_letter_key_var.set("")
        self.offer_letter_entry.configure(placeholder_text="No file selected")

    def store_offer_letter_in_db(self, pdf_key, file_path, student_name, company_name):
        """Store offer letter PDF in separate database using PDF manager"""
        try:
            # Use the PDF manager to store the PDF
            stored_pdf_key = pdf_manager.store_pdf(file_path, student_name, company_name)
            
            if stored_pdf_key:
                # Update the pdf_key variable to use the generated key
                self.offer_letter_key_var.set(stored_pdf_key)
                return True
            else:
                return False
                
        except Exception as e:
            messagebox.showerror("PDF Storage Error", f"Failed to store offer letter: {e}")
            return False

    def verify_offer_letter_exists(self, pdf_key):
        """Verify if offer letter exists in the separate database"""
        return pdf_manager.pdf_exists(pdf_key)

    def view_offer_letter(self, student_name, company_name):
        """View offer letter PDF for a specific student and company"""
        try:
            # Find the placement record to get PDF key
            placement = self.collection.find_one({
                "student_name": student_name,
                "company_name": company_name
            })
            
            if not placement or not placement.get("offer_letter_pdf_key"):
                messagebox.showwarning("No Offer Letter", "No offer letter found for this placement.")
                return
            
            pdf_key = placement["offer_letter_pdf_key"]
            
            # Use PDF manager to view the PDF
            result = pdf_manager.view_pdf(pdf_key)
            
            if result['success']:
                messagebox.showinfo("PDF Opened", result['message'])
            else:
                messagebox.showerror("Error Opening PDF", result['message'])
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to view offer letter: {e}")

    def add_placed_student(self):
        import os
        import datetime
        # Validate required fields
        if not all([self.student_name_var.get(), self.student_branch_var.get(),
                    self.placed_company_name_var.get(), self.placed_email_var.get(),
                    self.placed_contact_info_var.get(), self.placed_hr_name_var.get(),
                    self.placed_package_var.get(), self.year_of_placement_var.get(),
                    self.position_var.get(), self.batch_var.get()]):
            messagebox.showerror("Error", "Please fill all required fields (*)")
            return

        # Prepare placed student data
        student_name = self.student_name_var.get().upper()
        company_name = self.placed_company_name_var.get().upper()
        
        # Handle PDF storage using PDF manager
        pdf_key = None
        pdf_stored = False
        
        if self.offer_letter_path_var.get():
            try:
                # Use PDF manager to store the PDF
                pdf_key = pdf_manager.store_pdf(
                    self.offer_letter_path_var.get(),
                    student_name,
                    company_name
                )
                pdf_stored = True
                
            except Exception as e:
                result = messagebox.askyesno("PDF Storage Failed", 
                    f"Failed to store the offer letter PDF:\n{str(e)}\n\n"
                    "Do you want to continue saving the placement record without the PDF?")
                if not result:
                    return
                pdf_key = None
                pdf_stored = False
            
        placed_student_data = {
            "student_name": student_name,
            "student_branch": self.student_branch_var.get().upper(),
            "batch": self.batch_var.get().upper(),
            "company_name": company_name,
            "position": self.position_var.get().upper(),
            "year_of_placement": self.year_of_placement_var.get(),
            "package": self.placed_package_var.get().upper(),
            "email": self.placed_email_var.get(),  # Keep email in original format
            "contact_info": self.placed_contact_info_var.get(),
            "hr_name": self.placed_hr_name_var.get().upper(),
            "address": self.placed_address_var.get().upper(),
            # New fields
            "offer_letter_pdf_key": pdf_key,  # Key to link with offer letters database
            "placement_suggestion": self.placement_suggestion_var.get().upper(),
            "company_levels": self.company_levels_var.get().upper(),
            "skills_required": self.skills_required_var.get().upper(),
            "important_suggestions": self.important_suggestions_var.get().upper(),
            "created_date": datetime.datetime.now(),
            "has_offer_letter": pdf_stored
        }

        try:
            if self.collection is not None:
                # Check for duplicate
                existing = self.collection.find_one({
                    "$and": [
                        {"student_name": placed_student_data["student_name"]},
                        {"company_name": placed_student_data["company_name"]}
                    ]
                })

                if existing:
                    # Ask user if they want to update existing record
                    existing_student = existing.get("student_name", "Unknown")
                    existing_company = existing.get("company_name", "Unknown")
                    result = messagebox.askyesno("Duplicate Found", 
                        f"A placement record already exists:\n\n"
                        f"Student: {existing_student}\n"
                        f"Company: {existing_company}\n\n"
                        f"Do you want to UPDATE the existing record with new data?")
                    
                    if result:
                        # Update existing record
                        self.collection.update_one(
                            {"_id": existing["_id"]},
                            {"$set": placed_student_data}
                        )
                        # Invalidate cache after updating placement
                        invalidate_cache('placements')
                        success_msg = f"Placement record for '{existing_student}' updated successfully!"
                        if pdf_stored:
                            success_msg += f"\n📄 Offer letter stored with key: {pdf_key}"
                        messagebox.showinfo("Success", success_msg)
                        self.clear_placed_student_form()
                    return

                # Insert new record
                result = self.collection.insert_one(placed_student_data)
                if result.inserted_id:
                    # Invalidate cache after adding new placement
                    invalidate_cache('placements')
                    success_msg = "Placed student added successfully!"
                    if pdf_stored:
                        success_msg += f"\n📄 Offer letter stored with key: {pdf_key}"
                        success_msg += f"\n🔗 PDF linked to placement record"
                    messagebox.showinfo("Success", success_msg)
                    self.clear_placed_student_form()
                else:
                    messagebox.showerror("Error", "Failed to add placement record")
            else:
                messagebox.showerror("Error", "Database connection failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add placed student: {e}")

    def clear_placed_student_form(self):
        # Clear all form fields
        for var in [self.student_name_var, self.student_branch_var, self.placed_company_name_var,
                    self.placed_email_var, self.placed_contact_info_var, self.placed_hr_name_var,
                    self.placed_package_var, self.placed_address_var, self.batch_var,
                    self.position_var, self.year_of_placement_var, self.offer_letter_path_var,
                    self.placement_suggestion_var, self.company_levels_var, self.skills_required_var,
                    self.important_suggestions_var]:
            var.set("")
        
        # Clear file upload display
        if hasattr(self, 'offer_letter_entry'):
            self.offer_letter_entry.configure(placeholder_text="No file selected")
        if hasattr(self, 'offer_letter_key_var'):
            self.offer_letter_key_var.set("")
        
        # Clear textbox if it exists
        try:
            # Find and clear the textbox
            for widget in self.root.winfo_children():
                if hasattr(widget, 'winfo_children'):
                    for child in widget.winfo_children():
                        if isinstance(child, ctk.CTkTextbox):
                            child.delete("1.0", "end")
        except:
            pass

    def setup_view_placed_students_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="👁 VIEW PLACED STUDENTS",
                                   font=("Arial", 18, "bold"), text_color=COLORS["info"])
        title_label.pack(pady=5)

        scroll_frame = ctk.CTkFrame(tab, fg_color=COLORS["content_frame"])
        scroll_frame.pack(fill='both', expand=True, padx=5, pady=3)

        # Compact Search Frame
        search_frame = ctk.CTkFrame(scroll_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        search_frame.pack(fill='x', padx=5, pady=3)

        # Row 1: Student Name, Company, Branch
        search_row1 = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row1.pack(fill='x', padx=10, pady=4)

        ctk.CTkLabel(search_row1, text="Student:", font=("Arial", 11, "bold"), width=60).pack(side='left', padx=2)
        self.search_placed_student_name_var = ctk.StringVar()
        search_student_entry = ctk.CTkEntry(search_row1, textvariable=self.search_placed_student_name_var, width=150, height=32,
                     placeholder_text="Student name")
        search_student_entry.pack(side='left', padx=2)
        self.search_placed_student_name_var.trace_add('write', self.uppercase_entry(search_student_entry))

        ctk.CTkLabel(search_row1, text="Company:", font=("Arial", 11, "bold"), width=70).pack(side='left', padx=2)
        self.search_placed_company_var = ctk.StringVar()
        search_company_entry = ctk.CTkEntry(search_row1, textvariable=self.search_placed_company_var, width=150, height=32,
                     placeholder_text="Company name")
        search_company_entry.pack(side='left', padx=2)
        self.search_placed_company_var.trace_add('write', self.uppercase_entry(search_company_entry))

        ctk.CTkLabel(search_row1, text="Branch:", font=("Arial", 11, "bold"), width=55).pack(side='left', padx=2)
        self.search_placed_branch_var = ctk.StringVar()
        search_branch_entry = ctk.CTkEntry(search_row1, textvariable=self.search_placed_branch_var, width=100, height=32,
                     placeholder_text="CSE, ECE")
        search_branch_entry.pack(side='left', padx=2)
        self.search_placed_branch_var.trace_add('write', self.uppercase_entry(search_branch_entry))

        ctk.CTkLabel(search_row1, text="Package:", font=("Arial", 11, "bold"), width=60).pack(side='left', padx=2)
        self.search_placed_package_var = ctk.StringVar()
        ctk.CTkEntry(search_row1, textvariable=self.search_placed_package_var, width=100, height=32,
                     placeholder_text="Min ₹").pack(side='left', padx=2)

        ctk.CTkLabel(search_row1, text="HR:", font=("Arial", 11, "bold"), width=30).pack(side='left', padx=2)
        self.search_placed_hr_var = ctk.StringVar()
        search_hr_entry = ctk.CTkEntry(search_row1, textvariable=self.search_placed_hr_var, width=100, height=32,
                     placeholder_text="HR name")
        search_hr_entry.pack(side='left', padx=2)
        self.search_placed_hr_var.trace_add('write', self.uppercase_entry(search_hr_entry))

        # Row 2: Result Limit and Action Buttons
        search_row2 = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row2.pack(fill='x', padx=10, pady=4)

        ctk.CTkLabel(search_row2, text="Show:", font=("Arial", 11, "bold")).pack(side='left', padx=5)
        self.placed_result_limit_var = ctk.StringVar(value="50")
        limit_options = ["10", "20", "50", "100", "All"]
        ctk.CTkOptionMenu(search_row2, variable=self.placed_result_limit_var, values=limit_options,
                          width=80, height=32).pack(side='left', padx=3)

        ctk.CTkButton(search_row2, text="🔍 SEARCH", command=self.advanced_search_placed_students,
                      height=32, width=100, font=("Arial", 11, "bold"),
                      fg_color=COLORS["info"]).pack(side='left', padx=5)

        ctk.CTkButton(search_row2, text="🔄 CLEAR", command=self.clear_placed_filters,
                      height=32, width=80, font=("Arial", 11, "bold"),
                      fg_color=COLORS["warning"]).pack(side='left', padx=3)

        ctk.CTkButton(search_row2, text="📊 Export Excel", command=self.export_placed_students_excel,
                      height=32, width=110, font=("Arial", 11, "bold"),
                      fg_color=COLORS["success"]).pack(side='left', padx=3)

        ctk.CTkButton(search_row2, text="📄 Export CSV", command=self.export_placed_students_csv,
                      height=32, width=100, font=("Arial", 11, "bold"),
                      fg_color=COLORS["success"]).pack(side='left', padx=3)

        # Results frame - regular frame since table has its own scrolling
        self.placed_results_frame = ctk.CTkFrame(scroll_frame, fg_color=COLORS["content_frame"])
        self.placed_results_frame.pack(fill='both', expand=True, padx=5, pady=3)

        # Show instruction message
        instruction_label = ctk.CTkLabel(self.placed_results_frame,
                                         text="💡 Enter search criteria and click SEARCH",
                                         font=("Arial", 12),
                                         text_color=COLORS["info"])
        instruction_label.pack(pady=30)

    def clear_placed_filters(self):
        """Clear all search filters using optimized batch clearing"""
        search_vars = [
            self.search_placed_student_name_var, self.search_placed_company_var,
            self.search_placed_branch_var, self.search_placed_package_var, self.search_placed_hr_var
        ]
        
        for var in search_vars:
            var.set("")
        
        self.placed_result_limit_var.set("50")

        # Efficient widget clearing
        batch_clear_widgets(self.placed_results_frame)

        instruction_label = ctk.CTkLabel(self.placed_results_frame,
                                         text="💡 Enter search criteria above and click SEARCH",
                                         font=("Arial", 13),
                                         text_color=COLORS["info"])
        instruction_label.pack(pady=50)

    def advanced_search_placed_students(self):
        """Optimized advanced search with multiple filters"""
        batch_clear_widgets(self.placed_results_frame)

        # Show loading indicator
        loading_label = ctk.CTkLabel(self.placed_results_frame, text="⏳ Searching placements...",
                                     font=("Arial", 16, "bold"), text_color=COLORS["info"])
        loading_label.pack(pady=50)
        self.placed_results_frame.update()

        try:
            if self.collection is None:
                loading_label.destroy()
                ctk.CTkLabel(self.placed_results_frame, text="Database connection failed",
                             font=("Arial", 14)).pack(pady=20)
                return

            # Build optimized query using utility function
            filters = {
                "student_name": self.search_placed_student_name_var.get().strip(),
                "company_name": self.search_placed_company_var.get().strip(),
                "student_branch": self.search_placed_branch_var.get().strip(),
                "hr_name": self.search_placed_hr_var.get().strip()
            }
            
            query_conditions = []
            for field, value in filters.items():
                if value:
                    query_conditions.append({field: {"$regex": value, "$options": "i"}})

            # Package filter with optimized numeric extraction
            package = self.search_placed_package_var.get().strip()
            if package:
                package_num = extract_numeric_value(package)
                if package_num > 0:
                    query_conditions.append({"package": {"$regex": str(int(package_num)), "$options": "i"}})
                else:
                    query_conditions.append({"package": {"$regex": package, "$options": "i"}})

            query = {"$and": query_conditions} if query_conditions else {}

            # Get result limit efficiently
            limit_str = self.placed_result_limit_var.get()
            limit = None if limit_str == "All" else safe_int_convert(limit_str, 50)

            # Optimized database query with projection - Include ALL fields for new functionality
            projection = {
                "student_name": 1, "student_branch": 1, "batch": 1, "company_name": 1,
                "position": 1, "year_of_placement": 1, "package": 1, "hr_name": 1,
                "contact_info": 1, "email": 1, "address": 1,
                # Include new fields
                "offer_letter_pdf_key": 1, "placement_suggestion": 1, "company_levels": 1,
                "skills_required": 1, "important_suggestions": 1, "has_offer_letter": 1,
                "created_date": 1
            }
            
            cursor = self.collection.find(query, projection).sort("_id", -1)
            if limit:
                cursor = cursor.limit(limit)
            
            placed_students = list(cursor)

            # Client-side package filtering if numeric threshold provided
            if package:
                package_threshold = extract_numeric_value(package)
                if package_threshold > 0:
                    placed_students = self._filter_placements_by_package(placed_students, package_threshold)

            loading_label.destroy()

            if not placed_students:
                ctk.CTkLabel(self.placed_results_frame, text="❌ No placements found matching your criteria",
                             font=("Arial", 14), text_color=COLORS["error"]).pack(pady=20)
                return

            # Format filter info efficiently
            filter_text = format_filter_info(filters)
            if package:
                filter_text += f" | Package: {package}"

            # Store current placements for export
            self.current_placements = placed_students

            # Create professional table
            self.create_professional_placed_table(placed_students, filter_text)

        except Exception as e:
            if 'loading_label' in locals():
                loading_label.destroy()
            ctk.CTkLabel(self.placed_results_frame, text=f"Error searching placements: {e}",
                         font=("Arial", 14), text_color=COLORS["error"]).pack(pady=20)

    def _filter_placements_by_package(self, placements, threshold):
        """Efficiently filter placements by package threshold - O(n) single pass"""
        filtered = []
        for placed in placements:
            placed_package = placed.get("package", "")
            if placed_package:
                placed_package_num = extract_numeric_value(placed_package)
                # Convert LPA to actual amount if needed
                if "lpa" in str(placed_package).lower():
                    placed_package_num *= 100000
                if placed_package_num >= threshold:
                    filtered.append(placed)
            else:
                filtered.append(placed)  # Include placements without package info
        return filtered

    def create_professional_placed_table(self, placements, filter_text=""):
        """Create a high-performance professional table using ttk.Treeview for placements"""
        import tkinter as tk
        from tkinter import ttk

        # Create table container
        table_container = ctk.CTkFrame(self.placed_results_frame, fg_color=COLORS["content_frame"])
        table_container.pack(fill='both', expand=True, padx=5, pady=5)

        # Configure ttk style for dark theme
        style = ttk.Style()
        style.theme_use('clam')

        # Configure Treeview colors
        style.configure("Placed.Treeview",
            background="#1e1e1e",
            foreground="white",
            fieldbackground="#1e1e1e",
            rowheight=35,
            font=("Arial", 11)
        )
        style.configure("Placed.Treeview.Heading",
            background="#3498DB",
            foreground="white",
            font=("Arial", 11, "bold"),
            padding=8
        )
        style.map("Placed.Treeview",
            background=[("selected", "#3498DB")],
            foreground=[("selected", "white")]
        )
        style.map("Placed.Treeview.Heading",
            background=[("active", "#2980b9")]
        )

        # Define columns with new fields
        columns = ("idx", "student_name", "branch", "batch", "company", "position", "year", "package", "hr_name", "contact", "email", "address", "offer_letter", "suggestion", "levels", "skills", "notes")

        # Create Treeview with scrollbars
        tree_frame = tk.Frame(table_container, bg="#1e1e1e")
        tree_frame.pack(fill='both', expand=True)

        # Scrollbars
        v_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
        v_scroll.pack(side="right", fill="y")

        h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        h_scroll.pack(side="bottom", fill="x")

        # Create Treeview
        self.placed_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            style="Placed.Treeview",
            yscrollcommand=v_scroll.set,
            xscrollcommand=h_scroll.set
        )

        v_scroll.config(command=self.placed_tree.yview)
        h_scroll.config(command=self.placed_tree.xview)

        # Define column headings and widths with new fields
        col_config = [
            ("idx", "#", 50),
            ("student_name", "Student Name", 180),
            ("branch", "Branch", 100),
            ("batch", "Batch", 80),
            ("company", "Company", 180),
            ("position", "Position", 150),
            ("year", "Year", 80),
            ("package", "Package", 100),
            ("hr_name", "HR Name", 150),
            ("contact", "Contact", 120),
            ("email", "Email", 200),
            ("address", "Address", 200),
            ("offer_letter", "Offer Letter", 120),
            ("suggestion", "Placement Suggestion", 200),
            ("levels", "Company Levels", 120),
            ("skills", "Skills Required", 200),
            ("notes", "Important Notes", 200)
        ]

        for col_id, heading, width in col_config:
            self.placed_tree.heading(col_id, text=heading, anchor="center")
            self.placed_tree.column(col_id, width=width, minwidth=width, anchor="center")

        self.placed_tree.pack(fill='both', expand=True)

        # Configure row tags for alternating colors
        self.placed_tree.tag_configure("oddrow", background="#2b2b2b")
        self.placed_tree.tag_configure("evenrow", background="#1e1e1e")
        self.placed_tree.tag_configure("high_package", foreground="#2CC985")
        self.placed_tree.tag_configure("medium_package", foreground="#F39C12")

        # Insert data
        for idx, placed in enumerate(placements, 1):
            try:
                student_name = placed.get("student_name", "N/A")
                branch = placed.get("student_branch", "N/A")
                batch = placed.get("batch", "N/A")
                company = placed.get("company_name", "N/A")
                position = placed.get("position", "N/A")
                year = placed.get("year_of_placement", "N/A")
                package = placed.get("package", "N/A")
                hr_name = placed.get("hr_name", "N/A")
                contact = placed.get("contact_info", "N/A")
                email = placed.get("email", "N/A")
                address = placed.get("address", "N/A")
                
                # New fields - Enhanced offer letter handling
                pdf_key = placed.get("offer_letter_pdf_key")
                offer_letter = "No"
                
                if pdf_key:
                    # Verify PDF exists in offer letters database using PDF manager
                    if self.verify_offer_letter_exists(pdf_key):
                        offer_letter = "📄 Click to View"
                    else:
                        offer_letter = "❌ Missing"
                
                suggestion = placed.get("placement_suggestion", "N/A")
                levels = placed.get("company_levels", "N/A")
                skills = placed.get("skills_required", "N/A")
                notes = placed.get("important_suggestions", "N/A")

                values = (idx, student_name, branch, batch, company, position, year, package, hr_name, contact, email, address, offer_letter, suggestion, levels, skills, notes)

                # Determine row tags
                tags = ["oddrow"] if idx % 2 == 1 else ["evenrow"]

                # Color code by package
                try:
                    import re
                    numbers = re.findall(r'\d+\.?\d*', str(package))
                    if numbers:
                        pkg_val = float(numbers[0])
                        if "lpa" in str(package).lower():
                            pkg_val = pkg_val * 100000
                        if pkg_val >= 1000000:  # >= 10 LPA
                            tags.append("high_package")
                        elif pkg_val >= 500000:  # >= 5 LPA
                            tags.append("medium_package")
                except:
                    pass

                self.placed_tree.insert("", "end", values=values, tags=tags)

            except Exception as e:
                pass  # Skip problematic records silently

        # Mouse wheel scrolling
        def on_mousewheel(event):
            self.placed_tree.yview_scroll(int(-1*(event.delta/120)), "units")

        self.placed_tree.bind("<MouseWheel>", on_mousewheel)
        
        # Double-click event to view offer letter
        def on_double_click(event):
            try:
                selection = self.placed_tree.selection()
                if not selection:
                    return
                
                item = selection[0]
                values = self.placed_tree.item(item, 'values')
                
                if len(values) > 12:  # Offer letter column index
                    offer_letter_status = values[12]  # Offer letter column
                    
                    # Check if it's a clickable PDF
                    if "📄 Click to View" in offer_letter_status or "📄 View PDF" in offer_letter_status:
                        student_name = values[1]  # Student name column
                        company_name = values[4]  # Company column
                        
                        self.view_offer_letter(student_name, company_name)
                    elif "❌ Missing" in offer_letter_status:
                        messagebox.showwarning("PDF Missing", "The offer letter PDF is missing from the database.")
                    elif offer_letter_status == "No":
                        messagebox.showinfo("No PDF", "No offer letter uploaded for this placement.")
                        
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open offer letter: {e}")
        
        self.placed_tree.bind("<Double-1>", on_double_click)
        
        # Add right-click context menu for additional options
        def on_right_click(event):
            try:
                # Select the item under cursor
                item = self.placed_tree.identify_row(event.y)
                if item:
                    self.placed_tree.selection_set(item)
                    
                    # Create context menu
                    context_menu = ctk.CTkToplevel()
                    context_menu.title("Options")
                    context_menu.geometry("200x100")
                    context_menu.transient(self.placed_tree.winfo_toplevel())
                    
                    values = self.placed_tree.item(item, 'values')
                    if len(values) > 12:
                        offer_letter_status = values[12]
                        student_name = values[1]
                        company_name = values[4]
                        
                        if "📄" in offer_letter_status:
                            ctk.CTkButton(context_menu, text="📄 View Offer Letter",
                                        command=lambda: [self.view_offer_letter(student_name, company_name), context_menu.destroy()]).pack(pady=5)
                        
                        ctk.CTkButton(context_menu, text="📋 Copy Student Name",
                                    command=lambda: [self.placed_tree.clipboard_clear(), self.placed_tree.clipboard_append(student_name), context_menu.destroy()]).pack(pady=5)
                        
                        ctk.CTkButton(context_menu, text="❌ Close",
                                    command=context_menu.destroy).pack(pady=5)
                    
                    # Position menu near cursor
                    context_menu.geometry(f"+{event.x_root}+{event.y_root}")
                    
            except Exception as e:
                pass  # Silently handle right-click errors
        
        self.placed_tree.bind("<Button-3>", on_right_click)  # Right-click

        # Info bar at bottom
        info_frame = ctk.CTkFrame(self.placed_results_frame, fg_color=COLORS["section_frame"], height=40)
        info_frame.pack(fill='x', padx=5, pady=3)
        info_frame.pack_propagate(False)

        ctk.CTkLabel(
            info_frame,
            text=f"📊 Found {len(placements)} placements | Filters: {filter_text} | Scroll with mouse wheel",
            font=("Arial", 10, "bold"),
            text_color=COLORS["success"]
        ).pack(side='left', padx=10, pady=8)

    def setup_edit_placed_student_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="✏️ EDIT PLACED STUDENT",
                                   font=("Arial", 18, "bold"), text_color=COLORS["warning"])
        title_label.pack(pady=5)

        # Compact search section - horizontal layout
        search_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        search_frame.pack(fill='x', padx=5, pady=3)

        search_row = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row.pack(fill='x', padx=10, pady=5)

        self.edit_placed_search_var = ctk.StringVar()
        self.edit_placed_company_var = ctk.StringVar()

        ctk.CTkLabel(search_row, text="Student:", font=("Arial", 11, "bold")).pack(side='left', padx=3)
        edit_name_entry = ctk.CTkEntry(search_row, textvariable=self.edit_placed_search_var, width=180, height=32,
                                       placeholder_text="Student name")
        edit_name_entry.pack(side='left', padx=3)
        self.edit_placed_search_var.trace_add('write', self.uppercase_entry(edit_name_entry))

        ctk.CTkLabel(search_row, text="Company:", font=("Arial", 11, "bold")).pack(side='left', padx=3)
        edit_company_entry = ctk.CTkEntry(search_row, textvariable=self.edit_placed_company_var, width=180, height=32,
                                          placeholder_text="Company name")
        edit_company_entry.pack(side='left', padx=3)
        self.edit_placed_company_var.trace_add('write', self.uppercase_entry(edit_company_entry))

        ctk.CTkButton(search_row, text="🔍 SEARCH", command=self.search_placed_student_for_edit,
                      fg_color=COLORS["info"], font=("Arial", 11, "bold"), height=32, width=100).pack(side='left', padx=5)

        # Edit form frame - regular frame that auto-adjusts to content
        self.edit_placed_form_frame = ctk.CTkFrame(tab, fg_color=COLORS["content_frame"])
        self.edit_placed_form_frame.pack(fill='both', expand=True, padx=5, pady=3)

    def search_placed_student_for_edit(self):
        student_name = self.edit_placed_search_var.get().strip().upper()
        company_name = self.edit_placed_company_var.get().strip().upper()

        if not student_name and not company_name:
            messagebox.showerror("Error", "Please enter student name or company name")
            return

        # Clear previous results
        for widget in self.edit_placed_form_frame.winfo_children():
            widget.destroy()

        try:
            if self.collection is not None:
                # Build query based on inputs
                if student_name and company_name:
                    # Both fields - more specific search
                    placements = list(self.collection.find({
                        "student_name": {"$regex": student_name, "$options": "i"},
                        "company_name": {"$regex": company_name, "$options": "i"}
                    }))
                elif student_name:
                    # Only student name - show all matching
                    placements = list(self.collection.find({
                        "student_name": {"$regex": student_name, "$options": "i"}
                    }))
                else:
                    # Only company name
                    placements = list(self.collection.find({
                        "company_name": {"$regex": company_name, "$options": "i"}
                    }))

                if not placements:
                    ctk.CTkLabel(self.edit_placed_form_frame, text="❌ No placement records found",
                                 font=("Arial", 12), text_color=COLORS["error"]).pack(pady=20)
                    return

                if len(placements) == 1:
                    # Single result - show edit form directly
                    self.show_edit_placed_student_form(placements[0])
                else:
                    # Multiple results - show list to select
                    ctk.CTkLabel(self.edit_placed_form_frame, 
                                 text=f"📋 Found {len(placements)} records - Click to edit:",
                                 font=("Arial", 12, "bold"), text_color=COLORS["info"]).pack(pady=5)

                    for placed in placements:
                        row = ctk.CTkFrame(self.edit_placed_form_frame, fg_color=COLORS["section_frame"], corner_radius=8)
                        row.pack(fill='x', padx=5, pady=2)

                        info = f"👤 {placed.get('student_name', 'N/A')}  |  🏢 {placed.get('company_name', 'N/A')}  |  📚 {placed.get('student_branch', 'N/A')}  |  💰 {placed.get('package', 'N/A')}"
                        ctk.CTkLabel(row, text=info, font=("Arial", 10)).pack(side='left', padx=10, pady=5)

                        ctk.CTkButton(row, text="✏️ EDIT", command=lambda p=placed: self.show_edit_placed_student_form(p),
                                      fg_color=COLORS["warning"], font=("Arial", 10, "bold"), 
                                      height=28, width=70).pack(side='right', padx=5, pady=3)
            else:
                messagebox.showerror("Error", "Database connection failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to search placed student: {e}")

    def show_edit_placed_student_form(self, placed_student):
        # Clear previous form
        for widget in self.edit_placed_form_frame.winfo_children():
            widget.destroy()

        self.current_edit_placed_id = placed_student["_id"]

        # Create compact edit form
        placed_frame = ctk.CTkFrame(self.edit_placed_form_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        placed_frame.pack(fill='x', padx=5, pady=5)

        # Header with title and update button
        header_row = ctk.CTkFrame(placed_frame, fg_color=COLORS["section_frame"])
        header_row.pack(fill='x', padx=10, pady=5)

        ctk.CTkLabel(header_row, text="📝 EDIT PLACEMENT RECORD",
                     font=("Arial", 14, "bold"), text_color=COLORS["warning"]).pack(side='left', padx=5)

        ctk.CTkButton(header_row, text="💾 UPDATE", command=self.update_placed_student,
                      fg_color=COLORS["warning"], font=("Arial", 11, "bold"), height=32, width=100).pack(side='right', padx=5)

        # Initialize variables with new fields
        self.edit_student_name_var = ctk.StringVar(value=placed_student["student_name"])
        self.edit_student_branch_var = ctk.StringVar(value=placed_student["student_branch"])
        self.edit_batch_var = ctk.StringVar(value=placed_student.get("batch", ""))
        self.edit_placed_company_name_var = ctk.StringVar(value=placed_student["company_name"])
        self.edit_position_var = ctk.StringVar(value=placed_student.get("position", ""))
        self.edit_year_of_placement_var = ctk.StringVar(value=placed_student.get("year_of_placement", ""))
        self.edit_placed_email_var = ctk.StringVar(value=placed_student["email"])
        self.edit_placed_contact_info_var = ctk.StringVar(value=placed_student["contact_info"])
        self.edit_placed_hr_name_var = ctk.StringVar(value=placed_student["hr_name"])
        self.edit_placed_package_var = ctk.StringVar(value=placed_student["package"])
        self.edit_placed_address_var = ctk.StringVar(value=placed_student.get("address", ""))
        # New fields for editing
        self.edit_offer_letter_path_var = ctk.StringVar(value=placed_student.get("offer_letter_path", ""))
        self.edit_placement_suggestion_var = ctk.StringVar(value=placed_student.get("placement_suggestion", ""))
        self.edit_company_levels_var = ctk.StringVar(value=placed_student.get("company_levels", ""))
        self.edit_skills_required_var = ctk.StringVar(value=placed_student.get("skills_required", ""))
        self.edit_important_suggestions_var = ctk.StringVar(value=placed_student.get("important_suggestions", ""))

        # Compact 2-column grid layout
        placed_grid = ctk.CTkFrame(placed_frame, fg_color=COLORS["section_frame"])
        placed_grid.pack(fill='x', padx=10, pady=5)

        # Row 1: Student Name, Branch
        ctk.CTkLabel(placed_grid, text="Student Name*:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=3, pady=4)
        name_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_student_name_var, width=200, height=32)
        name_entry.grid(row=0, column=1, sticky='w', padx=3, pady=4)
        self.edit_student_name_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(placed_grid, text="Branch*:", font=("Arial", 11, "bold")).grid(row=0, column=2, sticky='w', padx=3, pady=4)
        branch_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_student_branch_var, width=150, height=32)
        branch_entry.grid(row=0, column=3, sticky='w', padx=3, pady=4)
        self.edit_student_branch_var.trace_add('write', self.uppercase_entry(branch_entry))

        # Row 2: Batch, Company
        ctk.CTkLabel(placed_grid, text="Batch*:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky='w', padx=3, pady=4)
        batch_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_batch_var, width=200, height=32)
        batch_entry.grid(row=1, column=1, sticky='w', padx=3, pady=4)
        self.edit_batch_var.trace_add('write', self.uppercase_entry(batch_entry))

        ctk.CTkLabel(placed_grid, text="Company*:", font=("Arial", 11, "bold")).grid(row=1, column=2, sticky='w', padx=3, pady=4)
        company_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_placed_company_name_var, width=150, height=32)
        company_entry.grid(row=1, column=3, sticky='w', padx=3, pady=4)
        self.edit_placed_company_name_var.trace_add('write', self.uppercase_entry(company_entry))

        # Row 3: Position, Year of Placement
        ctk.CTkLabel(placed_grid, text="Position*:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', padx=3, pady=4)
        position_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_position_var, width=200, height=32)
        position_entry.grid(row=2, column=1, sticky='w', padx=3, pady=4)
        self.edit_position_var.trace_add('write', self.uppercase_entry(position_entry))

        ctk.CTkLabel(placed_grid, text="Year*:", font=("Arial", 11, "bold")).grid(row=2, column=2, sticky='w', padx=3, pady=4)
        year_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_year_of_placement_var, width=150, height=32)
        year_entry.grid(row=2, column=3, sticky='w', padx=3, pady=4)

        # Row 4: Package, Email
        ctk.CTkLabel(placed_grid, text="Package*:", font=("Arial", 11, "bold")).grid(row=3, column=0, sticky='w', padx=3, pady=4)
        package_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_placed_package_var, width=200, height=32)
        package_entry.grid(row=3, column=1, sticky='w', padx=3, pady=4)
        self.edit_placed_package_var.trace_add('write', self.uppercase_entry(package_entry))

        ctk.CTkLabel(placed_grid, text="Email*:", font=("Arial", 11, "bold")).grid(row=3, column=2, sticky='w', padx=3, pady=4)
        email_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_placed_email_var, width=150, height=32)
        email_entry.grid(row=3, column=3, sticky='w', padx=3, pady=4)

        # Row 5: Contact, HR Name
        ctk.CTkLabel(placed_grid, text="Contact*:", font=("Arial", 11, "bold")).grid(row=4, column=0, sticky='w', padx=3, pady=4)
        contact_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_placed_contact_info_var, width=200, height=32)
        contact_entry.grid(row=4, column=1, sticky='w', padx=3, pady=4)

        ctk.CTkLabel(placed_grid, text="HR Name*:", font=("Arial", 11, "bold")).grid(row=4, column=2, sticky='w', padx=3, pady=4)
        hr_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_placed_hr_name_var, width=150, height=32)
        hr_entry.grid(row=4, column=3, sticky='w', padx=3, pady=4)
        self.edit_placed_hr_name_var.trace_add('write', self.uppercase_entry(hr_entry))

        # Row 6: Address (full width)
        ctk.CTkLabel(placed_grid, text="Address:", font=("Arial", 11, "bold")).grid(row=5, column=0, sticky='w', padx=3, pady=4)
        address_entry = ctk.CTkEntry(placed_grid, textvariable=self.edit_placed_address_var, width=450, height=32)
        address_entry.grid(row=5, column=1, columnspan=3, sticky='w', padx=3, pady=4)
        self.edit_placed_address_var.trace_add('write', self.uppercase_entry(address_entry))

        # Additional Information Section
        additional_frame = ctk.CTkFrame(placed_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        additional_frame.pack(fill='x', padx=10, pady=10)

        ctk.CTkLabel(additional_frame, text="📄 ADDITIONAL INFORMATION",
                     font=("Arial", 12, "bold"), text_color=COLORS["info"]).pack(pady=5)

        additional_grid = ctk.CTkFrame(additional_frame, fg_color=COLORS["section_frame"])
        additional_grid.pack(fill='x', padx=10, pady=5)

        # Row 1: Offer Letter Status, Placement Suggestion
        ctk.CTkLabel(additional_grid, text="Offer Letter:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=3, pady=4)
        offer_status = "Yes" if self.edit_offer_letter_path_var.get() else "No"
        ctk.CTkLabel(additional_grid, text=offer_status, font=("Arial", 11), text_color=COLORS["success"] if offer_status == "Yes" else COLORS["error"]).grid(row=0, column=1, sticky='w', padx=3, pady=4)

        ctk.CTkLabel(additional_grid, text="Placement Suggestion:", font=("Arial", 11, "bold")).grid(row=0, column=2, sticky='w', padx=3, pady=4)
        suggestion_entry = ctk.CTkEntry(additional_grid, textvariable=self.edit_placement_suggestion_var, width=200, height=32)
        suggestion_entry.grid(row=0, column=3, sticky='w', padx=3, pady=4)
        self.edit_placement_suggestion_var.trace_add('write', self.uppercase_entry(suggestion_entry))

        # Row 2: Company Levels, Skills Required
        ctk.CTkLabel(additional_grid, text="Company Levels:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky='w', padx=3, pady=4)
        levels_entry = ctk.CTkEntry(additional_grid, textvariable=self.edit_company_levels_var, width=200, height=32)
        levels_entry.grid(row=1, column=1, sticky='w', padx=3, pady=4)
        self.edit_company_levels_var.trace_add('write', self.uppercase_entry(levels_entry))

        ctk.CTkLabel(additional_grid, text="Skills Required:", font=("Arial", 11, "bold")).grid(row=1, column=2, sticky='w', padx=3, pady=4)
        skills_entry = ctk.CTkEntry(additional_grid, textvariable=self.edit_skills_required_var, width=200, height=32)
        skills_entry.grid(row=1, column=3, sticky='w', padx=3, pady=4)
        self.edit_skills_required_var.trace_add('write', self.uppercase_entry(skills_entry))

        # Row 3: Important Suggestions (full width)
        ctk.CTkLabel(additional_grid, text="Important Suggestions:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', padx=3, pady=4)
        suggestions_entry = ctk.CTkEntry(additional_grid, textvariable=self.edit_important_suggestions_var, width=450, height=32)
        suggestions_entry.grid(row=2, column=1, columnspan=3, sticky='w', padx=3, pady=4)
        self.edit_important_suggestions_var.trace_add('write', self.uppercase_entry(suggestions_entry))

    def update_placed_student(self):
        # Validate required fields including new fields
        if not all([self.edit_student_name_var.get(), self.edit_student_branch_var.get(),
                    self.edit_placed_company_name_var.get(), self.edit_placed_email_var.get(),
                    self.edit_placed_contact_info_var.get(), self.edit_placed_hr_name_var.get(),
                    self.edit_placed_package_var.get(), self.edit_batch_var.get(),
                    self.edit_position_var.get(), self.edit_year_of_placement_var.get()]):
            messagebox.showerror("Error", "Please fill all required fields (*)")
            return

        # Prepare updated placed student data with new fields
        updated_data = {
            "student_name": self.edit_student_name_var.get().upper(),
            "student_branch": self.edit_student_branch_var.get().upper(),
            "batch": self.edit_batch_var.get().upper(),
            "company_name": self.edit_placed_company_name_var.get().upper(),
            "position": self.edit_position_var.get().upper(),
            "year_of_placement": self.edit_year_of_placement_var.get(),
            "email": self.edit_placed_email_var.get(),
            "contact_info": self.edit_placed_contact_info_var.get(),
            "hr_name": self.edit_placed_hr_name_var.get().upper(),
            "package": self.edit_placed_package_var.get().upper(),
            "address": self.edit_placed_address_var.get().upper(),
            # New fields
            "placement_suggestion": self.edit_placement_suggestion_var.get().upper(),
            "company_levels": self.edit_company_levels_var.get().upper(),
            "skills_required": self.edit_skills_required_var.get().upper(),
            "important_suggestions": self.edit_important_suggestions_var.get().upper()
        }

        try:
            if self.collection is not None:
                self.collection.update_one(
                    {"_id": self.current_edit_placed_id},
                    {"$set": updated_data}
                )
                messagebox.showinfo("Success", "Placed student record updated successfully!")
                # Clear edit form
                for widget in self.edit_placed_form_frame.winfo_children():
                    widget.destroy()
                self.edit_placed_search_var.set("")
                self.edit_placed_company_var.set("")
            else:
                messagebox.showerror("Error", "Database connection failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update placed student: {e}")

    def setup_delete_placed_student_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="🗑️ DELETE PLACED STUDENT",
                                   font=("Arial", 18, "bold"), text_color=COLORS["error"])
        title_label.pack(pady=5)

        # Compact search section - horizontal layout
        search_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        search_frame.pack(fill='x', padx=5, pady=3)

        search_row = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row.pack(fill='x', padx=10, pady=5)

        self.delete_placed_search_var = ctk.StringVar()
        self.delete_placed_company_var = ctk.StringVar()

        ctk.CTkLabel(search_row, text="Student:", font=("Arial", 11, "bold")).pack(side='left', padx=3)
        name_entry = ctk.CTkEntry(search_row, textvariable=self.delete_placed_search_var, width=180, height=32,
                                  placeholder_text="Student name")
        name_entry.pack(side='left', padx=3)
        self.delete_placed_search_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(search_row, text="Company:", font=("Arial", 11, "bold")).pack(side='left', padx=3)
        company_entry = ctk.CTkEntry(search_row, textvariable=self.delete_placed_company_var, width=180, height=32,
                                     placeholder_text="Company name")
        company_entry.pack(side='left', padx=3)
        self.delete_placed_company_var.trace_add('write', self.uppercase_entry(company_entry))

        ctk.CTkButton(search_row, text="🔍 SEARCH", command=self.search_placed_student_for_delete,
                      fg_color=COLORS["info"], font=("Arial", 11, "bold"), height=32, width=100).pack(side='left', padx=5)

        # Results frame
        self.delete_placed_results_frame = ctk.CTkScrollableFrame(tab, fg_color=COLORS["content_frame"])
        self.delete_placed_results_frame.pack(fill='both', expand=True, padx=5, pady=3)

    def search_placed_student_for_delete(self):
        """Search for placed student by name and/or company"""
        student_name = self.delete_placed_search_var.get().strip().upper()
        company_name = self.delete_placed_company_var.get().strip().upper()

        if not student_name and not company_name:
            messagebox.showerror("Error", "Please enter at least Student Name or Company Name")
            return

        try:
            if self.collection is not None:
                # Build query
                query_conditions = []
                if student_name:
                    query_conditions.append({"student_name": {"$regex": student_name, "$options": "i"}})
                if company_name:
                    query_conditions.append({"company_name": {"$regex": company_name, "$options": "i"}})

                if len(query_conditions) == 2:
                    query = {"$and": query_conditions}
                else:
                    query = query_conditions[0]

                placed_students = list(self.collection.find(query))

                # Clear previous results efficiently
                batch_clear_widgets(self.delete_placed_results_frame)

                if not placed_students:
                    ctk.CTkLabel(self.delete_placed_results_frame,
                                 text="❌ No placement records found matching criteria",
                                 font=("Arial", 14), text_color=COLORS["error"]).pack(pady=20)
                    return

                # Display found records - compact horizontal layout
                for placed in placed_students:
                    placed_card = ctk.CTkFrame(self.delete_placed_results_frame, fg_color=COLORS["section_frame"],
                                               corner_radius=8)
                    placed_card.pack(fill='x', padx=5, pady=3)

                    info_row = ctk.CTkFrame(placed_card, fg_color=COLORS["section_frame"])
                    info_row.pack(fill='x', padx=10, pady=5)

                    ctk.CTkLabel(info_row, text=f"👤 {placed['student_name']}", font=("Arial", 11, "bold"),
                                 width=150).pack(side='left', padx=3)
                    ctk.CTkLabel(info_row, text=f"🏢 {placed['company_name']}", font=("Arial", 11),
                                 width=150).pack(side='left', padx=3)
                    ctk.CTkLabel(info_row, text=f"📚 {placed['student_branch']}", font=("Arial", 11),
                                 width=80).pack(side='left', padx=3)
                    ctk.CTkLabel(info_row, text=f"💰 {placed['package']}", font=("Arial", 11),
                                 width=100).pack(side='left', padx=3)
                    ctk.CTkLabel(info_row, text=f"👔 {placed['hr_name']}", font=("Arial", 11),
                                 width=120).pack(side='left', padx=3)

                    ctk.CTkButton(info_row, text="🗑️ DELETE",
                                  command=lambda p=placed: self.confirm_delete_placed_student(p),
                                  fg_color=COLORS["error"], font=("Arial", 10, "bold"), height=28, width=80).pack(side='right', padx=5)

        except Exception as e:
            messagebox.showerror("Error", f"Search failed: {e}")

    def confirm_delete_placed_student(self, placed):
        """Confirm and delete placed student record"""
        if messagebox.askyesno("Confirm Delete",
                               f"Are you sure you want to delete:\n{placed['student_name']} at {placed['company_name']}?\n\nThis action cannot be undone!"):
            try:
                if self.collection is not None:
                    self.collection.delete_one({"_id": placed["_id"]})
                    messagebox.showinfo("Success", "Placement record deleted successfully!")
                    self.delete_placed_search_var.set("")
                    self.delete_placed_company_var.set("")
                    # Refresh results efficiently
                    batch_clear_widgets(self.delete_placed_results_frame)
                else:
                    messagebox.showerror("Error", "Database connection failed")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete record: {e}")

    def export_placed_students_excel(self):
        """Export placed students to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime

            # Get placements from stored data
            if not hasattr(self, 'current_placements') or not self.current_placements:
                messagebox.showwarning("No Data", "No placements to export. Please search first.")
                return

            placements = self.current_placements

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Placements"

            # Add headers
            headers = ["Student Name", "Branch", "Company", "Package", "HR Name", "Contact", "Email", "Address", "Offer Letter", "Placement Suggestion", "Company Levels", "Skills Required", "Important Notes"]
            ws.append(headers)

            # Style header
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data
            for placement in placements:
                pdf_key = placement.get("offer_letter_pdf_key")
                offer_letter_status = "No"
                if pdf_key and self.verify_offer_letter_exists(pdf_key):
                    offer_letter_status = "Yes"
                elif pdf_key:
                    offer_letter_status = "Missing"
                
                ws.append([
                    placement.get("student_name", ""),
                    placement.get("student_branch", ""),
                    placement.get("company_name", ""),
                    placement.get("package", ""),
                    placement.get("hr_name", ""),
                    placement.get("contact_info", ""),
                    placement.get("email", ""),
                    placement.get("address", ""),
                    offer_letter_status,
                    placement.get("placement_suggestion", ""),
                    placement.get("company_levels", ""),
                    placement.get("skills_required", ""),
                    placement.get("important_suggestions", "")
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 30
            ws.column_dimensions['M'].width = 30
            ws.column_dimensions['H'].width = 30

            # Open Save As dialog
            from tkinter import filedialog
            default_filename = f"placements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Placements As"
            )
            
            if not filename:  # User cancelled
                return
                
            wb.save(filename)
            messagebox.showinfo("Success", f"Placements exported to:\n{filename}")
        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed. Install it using: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {e}")

    def export_placed_students_csv(self):
        """Export placed students to CSV file"""
        try:
            import csv
            from datetime import datetime
            from tkinter import filedialog

            # Get placements from stored data
            if not hasattr(self, 'current_placements') or not self.current_placements:
                messagebox.showwarning("No Data", "No placements to export. Please search first.")
                return

            placements = self.current_placements

            # Open Save As dialog
            default_filename = f"placements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Placements As"
            )
            
            if not filename:  # User cancelled
                return

            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ["Student Name", "Branch", "Company", "Package", "HR Name", "Contact", "Email", "Address", 
                             "Offer Letter", "Placement Suggestion", "Company Levels", "Skills Required", "Important Notes"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for placement in placements:
                    pdf_key = placement.get("offer_letter_pdf_key")
                    offer_letter_status = "No"
                    if pdf_key and self.verify_offer_letter_exists(pdf_key):
                        offer_letter_status = "Yes"
                    elif pdf_key:
                        offer_letter_status = "Missing"
                    
                    writer.writerow({
                        "Student Name": placement.get("student_name", ""),
                        "Branch": placement.get("student_branch", ""),
                        "Company": placement.get("company_name", ""),
                        "Package": placement.get("package", ""),
                        "HR Name": placement.get("hr_name", ""),
                        "Contact": placement.get("contact_info", ""),
                        "Email": placement.get("email", ""),
                        "Address": placement.get("address", ""),
                        "Offer Letter": offer_letter_status,
                        "Placement Suggestion": placement.get("placement_suggestion", ""),
                        "Company Levels": placement.get("company_levels", ""),
                        "Skills Required": placement.get("skills_required", ""),
                        "Important Notes": placement.get("important_suggestions", "")
                    })

            messagebox.showinfo("Success", f"Placements exported to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to CSV: {e}")

    def export_placed_charts_to_excel(self):
        """Export placement chart data to Excel with charts"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            from datetime import datetime
            from tkinter import filedialog

            if not hasattr(self, 'current_chart_data') or not self.current_chart_data:
                messagebox.showwarning("No Data", "No chart data to export. Please load charts first.")
                return

            placements = self.current_chart_data.get('placements', [])
            if not placements:
                messagebox.showwarning("No Data", "No placement data to export.")
                return

            # Open Save As dialog
            default_filename = f"placement_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Placement Analytics As"
            )

            if not filename:  # User cancelled
                return

            # Create workbook
            wb = openpyxl.Workbook()

            # Styles
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # ========== Sheet 1: Placements by Branch (Pie Chart) ==========
            ws1 = wb.active
            ws1.title = "Placements by Branch"
            ws1.append(["Branch", "Number of Placements"])

            branch_labels, branch_counts = self.get_placements_by_branch(placements)
            for i in range(len(branch_labels)):
                ws1.append([branch_labels[i], branch_counts[i]])

            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 22

            # Create Pie Chart
            if len(branch_labels) > 0:
                pie_chart = PieChart()
                pie_chart.title = "Placements by Branch"
                data = Reference(ws1, min_col=2, min_row=1, max_row=len(branch_labels) + 1)
                cats = Reference(ws1, min_col=1, min_row=2, max_row=len(branch_labels) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(cats)
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.width = 15
                pie_chart.height = 10
                ws1.add_chart(pie_chart, "D2")

            # ========== Sheet 2: Placements by Package (Bar Chart) ==========
            ws2 = wb.create_sheet("Placements by Package")
            ws2.append(["Package", "Number of Placements"])

            package_labels, package_counts = self.get_placements_by_package(placements)
            for i in range(len(package_labels)):
                ws2.append([package_labels[i], package_counts[i]])

            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 22

            # Create Bar Chart
            if len(package_labels) > 0:
                bar_chart = BarChart()
                bar_chart.title = "Placements by Package"
                bar_chart.type = "col"
                bar_chart.style = 10
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(package_labels) + 1)
                cats = Reference(ws2, min_col=1, min_row=2, max_row=len(package_labels) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(cats)
                bar_chart.dataLabels = DataLabelList()
                bar_chart.dataLabels.showVal = True
                bar_chart.width = 15
                bar_chart.height = 10
                ws2.add_chart(bar_chart, "D2")

            # ========== Sheet 3: Top Companies (Bar Chart) ==========
            ws3 = wb.create_sheet("Top Companies")
            ws3.append(["Company Name", "Number of Placements"])

            company_names, company_counts = self.get_top_companies(placements)
            for i in range(len(company_names)):
                ws3.append([company_names[i], company_counts[i]])

            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws3.column_dimensions['A'].width = 30
            ws3.column_dimensions['B'].width = 22

            # Create Bar Chart
            if len(company_names) > 0:
                bar_chart2 = BarChart()
                bar_chart2.title = "Top 10 Companies by Placements"
                bar_chart2.type = "col"
                bar_chart2.style = 10
                data = Reference(ws3, min_col=2, min_row=1, max_row=min(len(company_names) + 1, 11))
                cats = Reference(ws3, min_col=1, min_row=2, max_row=min(len(company_names) + 1, 11))
                bar_chart2.add_data(data, titles_from_data=True)
                bar_chart2.set_categories(cats)
                bar_chart2.dataLabels = DataLabelList()
                bar_chart2.dataLabels.showVal = True
                bar_chart2.width = 15
                bar_chart2.height = 10
                ws3.add_chart(bar_chart2, "D2")

            # ========== Sheet 4: Placements by HR (Pie Chart) ==========
            ws4 = wb.create_sheet("Placements by HR")
            ws4.append(["HR Name", "Number of Placements"])

            hr_names, hr_counts = self.get_placements_by_hr(placements)
            for i in range(len(hr_names)):
                ws4.append([hr_names[i], hr_counts[i]])

            for cell in ws4[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws4.column_dimensions['A'].width = 25
            ws4.column_dimensions['B'].width = 22

            # Create Pie Chart
            if len(hr_names) > 0:
                pie_chart2 = PieChart()
                pie_chart2.title = "Placements by HR"
                data = Reference(ws4, min_col=2, min_row=1, max_row=len(hr_names) + 1)
                cats = Reference(ws4, min_col=1, min_row=2, max_row=len(hr_names) + 1)
                pie_chart2.add_data(data, titles_from_data=True)
                pie_chart2.set_categories(cats)
                pie_chart2.dataLabels = DataLabelList()
                pie_chart2.dataLabels.showPercent = True
                pie_chart2.width = 15
                pie_chart2.height = 10
                ws4.add_chart(pie_chart2, "D2")

            # Save file
            wb.save(filename)
            
            # Count charts added
            chart_count = 0
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                chart_count += len(ws._charts)
            
            messagebox.showinfo("Success", 
                f"Placement analytics exported successfully to:\n{filename}\n\n"
                f"📊 Charts added: {chart_count}\n"
                f"📋 Sheets: {len(wb.sheetnames)}")

        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed. Install it using: pip install openpyxl")
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("Error", f"Failed to export to Excel: {e}\n\nDetails:\n{error_details[:500]}")
