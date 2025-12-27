import customtkinter as ctk
from tkinter import messagebox
from database_config import get_company_collection
from utils import (
    validate_email, validate_phone, uppercase_entry_handler,
    build_search_query, batch_clear_widgets, safe_int_convert, 
    safe_float_convert, format_filter_info, create_autopct_function,
    extract_numeric_value, COLORS, CHART_COLORS, get_cached_companies,
    invalidate_cache, get_matplotlib, performance_monitor
)


class CompanyManager:
    def __init__(self):
        self.collection = self.connect_db()

    def connect_db(self):
        return get_company_collection()

    def uppercase_entry(self, entry_widget):
        return uppercase_entry_handler(entry_widget)

    def show_add_interface(self, parent):
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_add_company_tab(parent)

    def show_view_interface(self, parent):
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_view_companies_tab(parent)

    def show_edit_interface(self, parent):
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_edit_company_tab(parent)

    def show_delete_interface(self, parent):
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_delete_company_tab(parent)

    def show_chart_interface(self, parent):
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_company_charts_tab(parent)

    def setup_company_charts_tab(self, parent):
        """Setup company analytics charts"""
        # Title
        title = ctk.CTkLabel(parent, text="📊 COMPANY ANALYTICS",
                             font=("Arial", 32, "bold"), text_color=COLORS["info"])
        title.pack(pady=3)

        # Export button
        export_btn = ctk.CTkButton(parent, text="📊 EXPORT TO EXCEL",
                                   command=self.export_company_charts_to_excel,
                                   height=35, width=180, font=("Arial", 12, "bold"),
                                   fg_color=COLORS["success"])
        export_btn.pack(pady=10)

        # Create scrollable frame for charts
        charts_container = ctk.CTkFrame(parent, fg_color=COLORS["content_frame"])
        charts_container.pack(fill='both', expand=True, padx=5, pady=5)

        # Loading indicator
        loading_label = ctk.CTkLabel(charts_container, text="⏳ Loading analytics...",
                                     font=("Arial", 14), text_color=COLORS["info"])
        loading_label.pack(pady=50)
        charts_container.update()

        try:
            # Get company data with pagination and projection
            companies = []
            if self.collection is not None:
                try:
                    projection = {
                        "company_name": 1, "company_Name": 1, "sector": 1, 
                        "package": 1, "hr_name": 1, "email": 1, "contact_info": 1
                    }
                    companies = list(self.collection.find({}, projection)
                                   .sort([("_id", -1)])
                                   .limit(500))  # Limit for performance
                except Exception as e:
                    print(f"Error fetching companies: {e}")
                    companies = []

            # Remove loading label
            loading_label.destroy()

            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.figure import Figure

            # Row 1: 2 charts
            row1_frame = ctk.CTkFrame(charts_container, fg_color=COLORS["content_frame"])
            row1_frame.pack(fill='both', expand=True, padx=5, pady=5)

            # Chart 1: Industry Type Distribution (Pie)
            self.create_company_chart(row1_frame, "Industry Type Distribution",
                                      self.get_industry_distribution(companies), "pie", row1_frame)

            # Chart 2: Company Package Distribution (Bar)
            self.create_company_chart(row1_frame, "Company Package Distribution",
                                      self.get_company_package_distribution(companies), "bar", row1_frame)

            # Row 2: Top 20 Companies by Package
            row2_frame = ctk.CTkFrame(charts_container, fg_color=COLORS["content_frame"])
            row2_frame.pack(fill='both', expand=True, padx=5, pady=5)

            # Chart 3: Top 20 Companies by Package (Horizontal Bar)
            self.create_company_chart(row2_frame, "Top 20 Companies by Package",
                                      self.get_top_companies_by_package(companies), "hbar", row2_frame)

            # Store chart data for export
            self.current_chart_data = {
                'companies': companies
            }

        except Exception as e:
            loading_label.destroy()
            error_label = ctk.CTkLabel(charts_container, text=f"Error loading analytics: {e}",
                                       font=("Arial", 14), text_color=COLORS["error"])
            error_label.pack(pady=50)

    def create_company_chart(self, parent, title, data, chart_type, row_frame):
        """Create a company chart"""
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.figure import Figure

        frame = ctk.CTkFrame(row_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        frame.pack(side='left', fill='both', expand=True, padx=5, pady=5)

        # Title
        title_label = ctk.CTkLabel(frame, text=title, font=("Arial", 14, "bold"),
                                   text_color=COLORS["info"])
        title_label.pack(pady=10)

        try:
            # Use larger height for horizontal bar charts with many items
            if chart_type == "hbar":
                fig = Figure(figsize=(8, 8), dpi=100, facecolor='#2b2b2b', edgecolor='none')
                fig.subplots_adjust(left=0.25, right=0.92, top=0.92, bottom=0.08)
            else:
                fig = Figure(figsize=(6.5, 5), dpi=100, facecolor='#2b2b2b', edgecolor='none')
                fig.subplots_adjust(left=0.12, right=0.95, top=0.88, bottom=0.18)
            ax = fig.add_subplot(111)
            ax.set_facecolor('#1e1e1e')

            if chart_type == "pie" and data:
                labels, sizes = data
                # Custom autopct function using utility
                autopct_func = create_autopct_function(sizes)
                ax.pie(sizes, labels=labels, autopct=autopct_func, colors=CHART_COLORS["primary"][:len(labels)],
                       textprops={'color': 'white', 'fontsize': 9, 'weight': 'bold'})
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)

            elif chart_type == "bar" and data:
                labels, values = data
                colors = ['#3498DB', '#2CC985', '#F39C12', '#E74C3C', '#9B59B6']
                bars = ax.bar(range(len(labels)), values, color=colors[:len(labels)])

                # Add value labels in the middle of bars for better visibility
                for i, bar in enumerate(bars):
                    height = bar.get_height()
                    if height > 0:  # Only add label if bar has height
                        # Position text in the middle of the bar
                        y_pos = height / 2
                        x_pos = bar.get_x() + bar.get_width() / 2.
                        
                        # Adjust font size based on bar height for better fit
                        font_size = min(12, max(8, int(height / 5)))
                        
                        ax.text(x_pos, y_pos, f'{int(height)}',
                                ha='center', va='center', color='white', 
                                fontsize=font_size, weight='bold',
                                bbox=dict(boxstyle='round,pad=0.3', facecolor='black', alpha=0.8))

                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right', color='white', fontsize=11, weight='bold')
                ax.tick_params(axis='y', colors='white', labelsize=10)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='y', alpha=0.3, color='white')

            elif chart_type == "hbar" and data:
                labels, values = data
                colors = ['#3498DB', '#2CC985', '#F39C12', '#E74C3C', '#9B59B6'] * 4
                
                # Add spacing between bars by using height parameter
                bar_height = 0.6  # Smaller height = more spacing
                y_positions = range(len(labels))
                bars = ax.barh(y_positions, values, height=bar_height, color=colors[:len(labels)])

                # Add value labels at the end of horizontal bars
                for i, bar in enumerate(bars):
                    width = bar.get_width()
                    # Position text at the end of the bar
                    ax.text(width + 0.5, bar.get_y() + bar.get_height() / 2.,
                            f'{int(width)}',
                            ha='left', va='center', color='white', fontsize=10, weight='bold')

                ax.set_yticks(y_positions)
                ax.set_yticklabels(labels, color='white', fontsize=8)
                ax.tick_params(axis='x', colors='white', labelsize=9)
                ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='x', alpha=0.3, color='white')
                # Add some padding to the right for labels
                ax.set_xlim(0, max(values) * 1.15 if values else 10)

            canvas = FigureCanvasTkAgg(fig, master=frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=5, pady=5)

        except Exception as e:
            error_label = ctk.CTkLabel(frame, text=f"Error: {str(e)[:50]}",
                                       font=("Arial", 10), text_color=COLORS["error"])
            error_label.pack(pady=10)

    def get_industry_distribution(self, companies):
        """Get industry/sector distribution"""
        sectors = {}
        for company in companies:
            sector = company.get("sector", "Unknown")
            sectors[sector] = sectors.get(sector, 0) + 1

        if not sectors:
            return (["No Data"], [0])

        return (list(sectors.keys()), list(sectors.values()))

    def get_company_package_distribution(self, companies):
        """Get company package distribution"""
        packages = {}
        for company in companies:
            pkg = company.get("package", "Unknown")
            packages[pkg] = packages.get(pkg, 0) + 1

        if not packages:
            return (["No Data"], [0])

        # Sort by package value
        sorted_pkg = sorted(packages.items(), key=lambda x: x[1], reverse=True)[:8]
        return (list(dict(sorted_pkg).keys()), list(dict(sorted_pkg).values()))

    def get_top_companies_by_package(self, companies):
        """Get top 20 companies by package"""
        import re
        company_packages = []

        for company in companies:
            name = company.get("company_name", "Unknown")
            pkg = company.get("package", "0")

            # Extract numeric value from package
            numbers = re.findall(r'\d+', str(pkg))
            if numbers:
                try:
                    pkg_num = float(numbers[0])
                    company_packages.append((name, pkg_num))
                except:
                    pass

        if not company_packages:
            return (["No Data"], [0])

        # Sort by package and get top 20
        sorted_companies = sorted(company_packages, key=lambda x: x[1], reverse=True)[:20]
        names = [c[0] for c in sorted_companies]
        packages = [c[1] for c in sorted_companies]

        return (names, packages)

    def setup_add_company_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="➕ ADD NEW COMPANY",
                                   font=("Arial", 18, "bold"), text_color=COLORS["success"])
        title_label.pack(pady=5)

        # Company Information Section - Compact 2-column layout
        company_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        company_frame.pack(fill='x', padx=10, pady=5)

        ctk.CTkLabel(company_frame, text="🏢 COMPANY INFORMATION",
                     font=("Arial", 14, "bold"), text_color=COLORS["info"]).pack(pady=8)

        # Company info fields - 2 column grid
        company_grid = ctk.CTkFrame(company_frame, fg_color=COLORS["section_frame"])
        company_grid.pack(fill='x', padx=15, pady=8)

        self.company_name_var = ctk.StringVar()
        self.company_email_var = ctk.StringVar()
        self.contact_info_var = ctk.StringVar()
        self.hr_name_var = ctk.StringVar()
        self.package_var = ctk.StringVar()
        self.company_website_var = ctk.StringVar()
        self.address_var = ctk.StringVar()

        # Row 1: Company Name, Email
        ctk.CTkLabel(company_grid, text="Company Name*:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=4)
        name_entry = ctk.CTkEntry(company_grid, textvariable=self.company_name_var, width=200, height=32)
        name_entry.grid(row=0, column=1, sticky='w', padx=5, pady=4)
        self.company_name_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(company_grid, text="Email*:", font=("Arial", 11, "bold")).grid(row=0, column=2, sticky='w', padx=5, pady=4)
        ctk.CTkEntry(company_grid, textvariable=self.company_email_var, width=200, height=32).grid(row=0, column=3, sticky='w', padx=5, pady=4)

        # Row 2: Contact, HR Name
        ctk.CTkLabel(company_grid, text="Contact Info*:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky='w', padx=5, pady=4)
        ctk.CTkEntry(company_grid, textvariable=self.contact_info_var, width=200, height=32).grid(row=1, column=1, sticky='w', padx=5, pady=4)

        ctk.CTkLabel(company_grid, text="HR Name*:", font=("Arial", 11, "bold")).grid(row=1, column=2, sticky='w', padx=5, pady=4)
        hr_entry = ctk.CTkEntry(company_grid, textvariable=self.hr_name_var, width=200, height=32)
        hr_entry.grid(row=1, column=3, sticky='w', padx=5, pady=4)
        self.hr_name_var.trace_add('write', self.uppercase_entry(hr_entry))

        # Row 3: Package, Website
        ctk.CTkLabel(company_grid, text="Package*:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', padx=5, pady=4)
        package_entry = ctk.CTkEntry(company_grid, textvariable=self.package_var, width=200, height=32)
        package_entry.grid(row=2, column=1, sticky='w', padx=5, pady=4)
        self.package_var.trace_add('write', self.uppercase_entry(package_entry))

        ctk.CTkLabel(company_grid, text="Website:", font=("Arial", 11, "bold")).grid(row=2, column=2, sticky='w', padx=5, pady=4)
        ctk.CTkEntry(company_grid, textvariable=self.company_website_var, width=200, height=32).grid(row=2, column=3, sticky='w', padx=5, pady=4)

        # Row 4: Address (full width)
        ctk.CTkLabel(company_grid, text="Address:", font=("Arial", 11, "bold")).grid(row=3, column=0, sticky='w', padx=5, pady=4)
        address_entry = ctk.CTkEntry(company_grid, textvariable=self.address_var, width=500, height=32)
        address_entry.grid(row=3, column=1, columnspan=3, sticky='w', padx=5, pady=4)
        self.address_var.trace_add('write', self.uppercase_entry(address_entry))

        # Submit button - compact
        submit_btn = ctk.CTkButton(tab, text="➕ ADD COMPANY", command=self.add_company,
                                   fg_color=COLORS["success"], font=("Arial", 14, "bold"), height=40, width=180)
        submit_btn.pack(pady=15)

    def add_company(self):
        # Validate required fields using optimized validation
        if not all([self.company_name_var.get(), self.company_email_var.get(),
                    self.contact_info_var.get(), self.hr_name_var.get(), self.package_var.get()]):
            messagebox.showerror("Error", "Please fill all required fields (*)")
            return

        if not validate_email(self.company_email_var.get()):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        # Prepare company data
        company_data = {
            "company_name": self.company_name_var.get().upper(),
            "email": self.company_email_var.get(),  # Keep email in original format
            "contact_info": self.contact_info_var.get(),
            "hr_name": self.hr_name_var.get().upper(),
            "package": self.package_var.get().upper(),
            "website": self.company_website_var.get(),  # Keep website URL in original format
            "address": self.address_var.get().upper()
        }

        try:
            if self.collection is not None:
                # Check for duplicate
                existing = self.collection.find_one({
                    "$or": [
                        {"company_name": company_data["company_name"]},
                        {"email": company_data["email"]}
                    ]
                })

                if existing:
                    # Ask user if they want to update existing record
                    existing_name = existing.get("company_name", "Unknown")
                    result = messagebox.askyesno("Duplicate Found", 
                        f"A company with the same name or email already exists:\n\n"
                        f"Existing: {existing_name}\n"
                        f"New: {company_data['company_name']}\n\n"
                        f"Do you want to UPDATE the existing record with new data?")
                    
                    if result:
                        # Update existing record
                        self.collection.update_one(
                            {"_id": existing["_id"]},
                            {"$set": company_data}
                        )
                        # Invalidate cache after updating company
                        invalidate_cache('companies')
                        messagebox.showinfo("Success", f"Company '{existing_name}' updated successfully!")
                        self.clear_company_form()
                    return

                self.collection.insert_one(company_data)
                # Invalidate cache after adding new company
                invalidate_cache('companies')
                messagebox.showinfo("Success", "Company added successfully!")
                self.clear_company_form()
            else:
                messagebox.showerror("Error", "Database connection failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add company: {e}")

    def clear_company_form(self):
        # Clear all form fields
        for var in [self.company_name_var, self.company_email_var, self.contact_info_var,
                    self.hr_name_var, self.package_var, self.company_website_var, self.address_var]:
            var.set("")

    def setup_view_companies_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="👁 ADVANCED COMPANY SEARCH",
                                   font=("Arial", 28, "bold"), text_color=COLORS["info"])
        title_label.pack(pady=10)

        scroll_frame = ctk.CTkFrame(tab, fg_color=COLORS["content_frame"])
        scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Advanced Search Frame
        search_frame = ctk.CTkFrame(scroll_frame, fg_color=COLORS["section_frame"], corner_radius=15)
        search_frame.pack(fill='x', padx=5, pady=8)

        ctk.CTkLabel(search_frame, text="🔍 LIBRARY-STYLE SEARCH SYSTEM",
                     font=("Arial", 20, "bold"), text_color=COLORS["info"]).pack(pady=15)

        # Row 1: Basic Search
        search_row1 = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row1.pack(fill='x', padx=20, pady=10)

        ctk.CTkLabel(search_row1, text="Company Name:", font=("Arial", 12, "bold"), width=120).pack(side='left', padx=5)
        self.search_company_name_var = ctk.StringVar()
        search_company_entry = ctk.CTkEntry(search_row1, textvariable=self.search_company_name_var, width=200, height=35,
                     placeholder_text="Company name")
        search_company_entry.pack(side='left', padx=5)
        self.search_company_name_var.trace_add('write', self.uppercase_entry(search_company_entry))

        ctk.CTkLabel(search_row1, text="Email:", font=("Arial", 12, "bold"), width=80).pack(side='left', padx=5)
        self.search_company_email_var = ctk.StringVar()
        ctk.CTkEntry(search_row1, textvariable=self.search_company_email_var, width=200, height=35,
                     placeholder_text="Email address").pack(side='left', padx=5)
        # Email field - no uppercase conversion

        ctk.CTkLabel(search_row1, text="HR Name:", font=("Arial", 12, "bold"), width=100).pack(side='left', padx=5)
        self.search_hr_name_var = ctk.StringVar()
        search_hr_entry = ctk.CTkEntry(search_row1, textvariable=self.search_hr_name_var, width=180, height=35,
                     placeholder_text="HR name")
        search_hr_entry.pack(side='left', padx=5)
        self.search_hr_name_var.trace_add('write', self.uppercase_entry(search_hr_entry))

        # Row 2: Package Filter (Numeric)
        search_row2 = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row2.pack(fill='x', padx=20, pady=10)

        ctk.CTkLabel(search_row2, text="Min Package (₹):", font=("Arial", 12, "bold"), width=120).pack(side='left',
                                                                                                       padx=5)
        self.search_package_var = ctk.StringVar()
        package_entry = ctk.CTkEntry(search_row2, textvariable=self.search_package_var, width=150, height=35,
                                     placeholder_text="e.g., 200000")
        package_entry.pack(side='left', padx=5)
        ctk.CTkLabel(search_row2, text="(Shows packages ≥ this amount)", font=("Arial", 9),
                     text_color=COLORS["info"]).pack(side='left', padx=5)

        ctk.CTkLabel(search_row2, text="Contact:", font=("Arial", 12, "bold"), width=100).pack(side='left', padx=5)
        self.search_contact_var = ctk.StringVar()
        ctk.CTkEntry(search_row2, textvariable=self.search_contact_var, width=150, height=35,
                     placeholder_text="Contact number").pack(side='left', padx=5)
        # Contact is numeric - no uppercase needed

        # Row 3: Result Limit and Action Buttons
        search_row3 = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row3.pack(fill='x', padx=20, pady=15)

        ctk.CTkLabel(search_row3, text="Show Results:", font=("Arial", 12, "bold")).pack(side='left', padx=10)
        self.company_result_limit_var = ctk.StringVar(value="50")
        limit_options = ["10", "20", "50", "100", "All"]
        ctk.CTkOptionMenu(search_row3, variable=self.company_result_limit_var, values=limit_options,
                          width=100, height=35).pack(side='left', padx=5)

        ctk.CTkButton(search_row3, text="🔍 SEARCH", command=self.advanced_search_companies,
                      height=35, width=120, font=("Arial", 13, "bold"),
                      fg_color=COLORS["info"]).pack(side='left', padx=10)

        ctk.CTkButton(search_row3, text="🔄 CLEAR", command=self.clear_company_filters,
                      height=35, width=100, font=("Arial", 13, "bold"),
                      fg_color=COLORS["warning"]).pack(side='left', padx=5)

        # Export buttons
        ctk.CTkButton(search_row3, text="📊 EXPORT EXCEL", command=self.export_companies_excel,
                      height=35, width=140, font=("Arial", 13, "bold"),
                      fg_color=COLORS["success"]).pack(side='left', padx=5)

        ctk.CTkButton(search_row3, text="📄 EXPORT CSV", command=self.export_companies_csv,
                      height=35, width=130, font=("Arial", 13, "bold"),
                      fg_color=COLORS["success"]).pack(side='left', padx=5)

        # Results frame - regular frame since Treeview has its own scrolling
        self.company_results_frame = ctk.CTkFrame(scroll_frame, fg_color=COLORS["content_frame"])
        self.company_results_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Show instruction message
        instruction_label = ctk.CTkLabel(self.company_results_frame,
                                         text="💡 Enter search criteria above and click SEARCH\n"
                                              "• Search by single field or combine multiple filters\n"
                                              "• Package: Enter minimum amount (e.g., 200000 shows all ≥ 200000)\n"
                                              "• Leave fields empty to search all companies\n"
                                              "• Select result limit from dropdown",
                                         font=("Arial", 13),
                                         text_color=COLORS["info"],
                                         justify="left")
        instruction_label.pack(pady=50)

    def load_companies(self, search_term=None, search_type="name", limit=None):
        # Clear previous results
        for widget in self.company_results_frame.winfo_children():
            widget.destroy()

        try:
            if self.collection is not None:
                query = {}
                if search_term:
                    if search_type == "name":
                        # Search both old and new field names
                        query = {"$or": [
                            {"company_name": {"$regex": search_term, "$options": "i"}},
                            {"company_Name": {"$regex": search_term, "$options": "i"}}
                        ]}
                    elif search_type == "email":
                        query = {"email": {"$regex": search_term, "$options": "i"}}
                    elif search_type == "hr":
                        query = {"hr_name": {"$regex": search_term, "$options": "i"}}

                # Fetch companies with optional limit, sorted by most recent
                if limit:
                    companies = list(self.collection.find(query).sort("_id", -1).limit(limit))
                else:
                    companies = list(self.collection.find(query).sort("_id", -1))

                if not companies:
                    ctk.CTkLabel(self.company_results_frame, text="No companies found",
                                 font=("Arial", 14)).pack(pady=20)
                    return

                # Create table header with proper column widths
                header_frame = ctk.CTkFrame(self.company_results_frame, fg_color=COLORS["info"], corner_radius=5)
                header_frame.pack(fill='x', pady=5, padx=5)

                headers = ["Company Name", "Email", "Contact", "HR Name", "Package", "Address"]
                # Define column widths for better alignment
                column_widths = [200, 200, 120, 150, 100, 200]

                for i, (header, width) in enumerate(zip(headers, column_widths)):
                    ctk.CTkLabel(header_frame, text=header, font=("Arial", 11, "bold"),
                                 text_color="white", width=width, anchor="w").grid(row=0, column=i, padx=3, pady=8,
                                                                                   sticky='ew')
                    header_frame.grid_columnconfigure(i, minsize=width)

                # Display companies
                for company in companies:
                    company_frame = ctk.CTkFrame(self.company_results_frame, fg_color=COLORS["section_frame"],
                                                 corner_radius=5)
                    company_frame.pack(fill='x', pady=2, padx=5)

                    # Handle both old and new data structures
                    data = [
                        company.get("company_name", company.get("company_Name", "N/A")),
                        company.get("email", "N/A"),
                        company.get("contact_info", company.get("contact_no", "N/A")),
                        company.get("hr_name", "N/A"),
                        company.get("package", "N/A"),
                        company.get("address", "N/A")
                    ]

                    for i, (value, width) in enumerate(zip(data, column_widths)):
                        ctk.CTkLabel(company_frame, text=value, font=("Arial", 10),
                                     width=width, anchor="w").grid(row=0, column=i, padx=3, pady=5, sticky='ew')
                        company_frame.grid_columnconfigure(i, minsize=width)

            else:
                ctk.CTkLabel(self.company_results_frame, text="Database connection failed",
                             font=("Arial", 14)).pack(pady=20)
        except Exception as e:
            ctk.CTkLabel(self.company_results_frame, text=f"Error loading companies: {e}",
                         font=("Arial", 14)).pack(pady=20)

    def clear_company_filters(self):
        """Clear all search filters using optimized batch clearing"""
        search_vars = [
            self.search_company_name_var, self.search_company_email_var, 
            self.search_hr_name_var, self.search_package_var, self.search_contact_var
        ]
        
        for var in search_vars:
            var.set("")
        
        self.company_result_limit_var.set("50")

        # Efficient widget clearing
        batch_clear_widgets(self.company_results_frame)

        instruction_label = ctk.CTkLabel(self.company_results_frame,
                                         text="💡 Enter search criteria above and click SEARCH",
                                         font=("Arial", 13),
                                         text_color=COLORS["info"])
        instruction_label.pack(pady=50)

    def advanced_search_companies(self):
        """Optimized advanced search with multiple filters"""
        batch_clear_widgets(self.company_results_frame)

        # Show loading indicator
        loading_label = ctk.CTkLabel(self.company_results_frame, text="⏳ Searching companies...",
                                     font=("Arial", 16, "bold"), text_color=COLORS["info"])
        loading_label.pack(pady=50)
        self.company_results_frame.update()

        try:
            if self.collection is None:
                loading_label.destroy()
                ctk.CTkLabel(self.company_results_frame, text="Database connection failed",
                             font=("Arial", 14)).pack(pady=20)
                return

            # Build optimized query using utility function
            filters = {
                "company_name": self.search_company_name_var.get().strip(),
                "email": self.search_company_email_var.get().strip(),
                "hr_name": self.search_hr_name_var.get().strip(),
                "contact_info": self.search_contact_var.get().strip()
            }
            
            # Build base query
            query_conditions = []
            for field, value in filters.items():
                if value:
                    if field == "company_name":
                        # Handle both old and new field names
                        query_conditions.append({"$or": [
                            {"company_name": {"$regex": value, "$options": "i"}},
                            {"company_Name": {"$regex": value, "$options": "i"}}
                        ]})
                    elif field == "contact_info":
                        # Handle both old and new field names
                        query_conditions.append({"$or": [
                            {"contact_info": {"$regex": value, "$options": "i"}},
                            {"contact_no": {"$regex": value, "$options": "i"}}
                        ]})
                    else:
                        query_conditions.append({field: {"$regex": value, "$options": "i"}})

            # Package filter with optimized numeric extraction
            package = self.search_package_var.get().strip()
            if package:
                package_num = extract_numeric_value(package)
                if package_num > 0:
                    # Use regex to find packages with this number
                    query_conditions.append({"package": {"$regex": str(int(package_num)), "$options": "i"}})
                else:
                    query_conditions.append({"package": {"$regex": package, "$options": "i"}})

            query = {"$and": query_conditions} if query_conditions else {}

            # Get result limit efficiently
            limit_str = self.company_result_limit_var.get()
            limit = None if limit_str == "All" else safe_int_convert(limit_str, 50)

            # Optimized database query with projection
            projection = {
                "company_name": 1, "company_Name": 1, "email": 1, "contact_info": 1, 
                "contact_no": 1, "hr_name": 1, "package": 1, "website": 1, "address": 1
            }
            
            cursor = self.collection.find(query, projection).sort("_id", -1)
            if limit:
                cursor = cursor.limit(limit)
            
            companies = list(cursor)

            # Client-side package filtering if numeric threshold provided
            if package:
                package_threshold = extract_numeric_value(package)
                if package_threshold > 0:
                    companies = self._filter_companies_by_package(companies, package_threshold)

            loading_label.destroy()

            if not companies:
                ctk.CTkLabel(self.company_results_frame, text="❌ No companies found matching your criteria",
                             font=("Arial", 14), text_color=COLORS["error"]).pack(pady=20)
                return

            # Format filter info efficiently
            filter_text = format_filter_info(filters)
            if package:
                filter_text += f" | Package: {package}"

            # Store current companies for export
            self.current_companies = companies

            # Create professional table
            self.create_professional_company_table(companies, filter_text)

        except Exception as e:
            if 'loading_label' in locals():
                loading_label.destroy()
            ctk.CTkLabel(self.company_results_frame, text=f"Error searching companies: {e}",
                         font=("Arial", 14), text_color=COLORS["error"]).pack(pady=20)

    def _filter_companies_by_package(self, companies, threshold):
        """Efficiently filter companies by package threshold - O(n) single pass"""
        filtered = []
        for comp in companies:
            comp_package = comp.get("package", "")
            if comp_package:
                comp_package_num = extract_numeric_value(comp_package)
                # Convert LPA to actual amount if needed
                if "lpa" in str(comp_package).lower():
                    comp_package_num *= 100000
                if comp_package_num >= threshold:
                    filtered.append(comp)
            else:
                filtered.append(comp)  # Include companies without package info
        return filtered

    def create_professional_company_table(self, companies, filter_text=""):
        """Create a high-performance professional table using ttk.Treeview for companies"""
        import tkinter as tk
        from tkinter import ttk

        # Create table container
        table_container = ctk.CTkFrame(self.company_results_frame, fg_color=COLORS["content_frame"])
        table_container.pack(fill='both', expand=True, padx=5, pady=5)

        # Configure ttk style for dark theme
        style = ttk.Style()
        style.theme_use('clam')

        # Configure Treeview colors
        style.configure("Company.Treeview",
            background="#1e1e1e",
            foreground="white",
            fieldbackground="#1e1e1e",
            rowheight=35,
            font=("Arial", 11)
        )
        style.configure("Company.Treeview.Heading",
            background="#3498DB",
            foreground="white",
            font=("Arial", 11, "bold"),
            padding=8
        )
        style.map("Company.Treeview",
            background=[("selected", "#3498DB")],
            foreground=[("selected", "white")]
        )
        style.map("Company.Treeview.Heading",
            background=[("active", "#2980b9")]
        )

        # Define columns
        columns = ("idx", "company_name", "email", "contact", "hr_name", "package", "website", "address")

        # Store website URLs for click handling
        self.company_website_urls = {}

        # Create Treeview with scrollbars
        tree_frame = tk.Frame(table_container, bg="#1e1e1e")
        tree_frame.pack(fill='both', expand=True)

        # Scrollbars
        v_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
        v_scroll.pack(side="right", fill="y")

        h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        h_scroll.pack(side="bottom", fill="x")

        # Create Treeview
        self.company_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            style="Company.Treeview",
            yscrollcommand=v_scroll.set,
            xscrollcommand=h_scroll.set
        )

        v_scroll.config(command=self.company_tree.yview)
        h_scroll.config(command=self.company_tree.xview)

        # Define column headings and widths
        col_config = [
            ("idx", "#", 50),
            ("company_name", "Company Name", 180),
            ("email", "Email", 200),
            ("contact", "Contact", 110),
            ("hr_name", "HR Name", 140),
            ("package", "Package", 100),
            ("website", "Website", 100),
            ("address", "Address", 200)
        ]

        for col_id, heading, width in col_config:
            self.company_tree.heading(col_id, text=heading, anchor="center")
            self.company_tree.column(col_id, width=width, minwidth=width, anchor="center")

        self.company_tree.pack(fill='both', expand=True)

        # Configure row tags for alternating colors
        self.company_tree.tag_configure("oddrow", background="#2b2b2b")
        self.company_tree.tag_configure("evenrow", background="#1e1e1e")
        self.company_tree.tag_configure("high_package", foreground="#2CC985")
        self.company_tree.tag_configure("medium_package", foreground="#F39C12")

        # Insert data
        import webbrowser
        for idx, company in enumerate(companies, 1):
            try:
                # Handle both old and new data structures
                company_name = company.get("company_name", company.get("company_Name", "N/A"))
                email = company.get("email", "N/A")
                contact = company.get("contact_info", company.get("contact_no", "N/A"))
                hr_name = company.get("hr_name", "N/A")
                package = company.get("package", "N/A")
                website_url = company.get("website", "")
                website_display = "🔗 Click" if website_url and website_url.strip() else "N/A"
                address = company.get("address", "N/A")

                values = (idx, company_name, email, contact, hr_name, package, website_display, address)

                # Determine row tags
                tags = ["oddrow"] if idx % 2 == 1 else ["evenrow"]

                # Color code by package (try to extract numeric value)
                try:
                    import re
                    numbers = re.findall(r'\d+\.?\d*', str(package))
                    if numbers:
                        pkg_val = float(numbers[0])
                        if "lpa" in str(package).lower():
                            pkg_val = pkg_val * 100000
                        if pkg_val >= 1000000:  # >= 10 LPA
                            tags.append("high_package")
                        elif pkg_val >= 500000:  # >= 5 LPA
                            tags.append("medium_package")
                except:
                    pass

                item_id = self.company_tree.insert("", "end", values=values, tags=tags)

                # Store website URL for this row
                if website_url and website_url.strip():
                    self.company_website_urls[item_id] = website_url

            except Exception as e:
                print(f"Error processing company {idx}: {e}")

        # Bind click on Website column to open URL
        def on_click(event):
            region = self.company_tree.identify_region(event.x, event.y)
            if region == "cell":
                item = self.company_tree.identify_row(event.y)
                column = self.company_tree.identify_column(event.x)
                if column == "#7" and item in self.company_website_urls:  # Website column
                    webbrowser.open(self.company_website_urls[item])

        self.company_tree.bind("<Button-1>", on_click)

        # Mouse wheel scrolling
        def on_mousewheel(event):
            self.company_tree.yview_scroll(int(-1*(event.delta/120)), "units")

        self.company_tree.bind("<MouseWheel>", on_mousewheel)

        # Info bar at bottom
        info_frame = ctk.CTkFrame(self.company_results_frame, fg_color=COLORS["section_frame"], height=40)
        info_frame.pack(fill='x', padx=5, pady=3)
        info_frame.pack_propagate(False)

        ctk.CTkLabel(
            info_frame,
            text=f"📊 Found {len(companies)} companies | Filters: {filter_text} | Scroll with mouse wheel",
            font=("Arial", 10, "bold"),
            text_color=COLORS["success"]
        ).pack(side='left', padx=10, pady=8)

    def setup_edit_company_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="✏️ EDIT COMPANY",
                                   font=("Arial", 18, "bold"), text_color=COLORS["warning"])
        title_label.pack(pady=5)

        # Two-field search section - horizontal layout
        search_frame = ctk.CTkFrame(tab, fg_color=COLORS["section_frame"], corner_radius=10)
        search_frame.pack(fill='x', padx=5, pady=3)

        search_row = ctk.CTkFrame(search_frame, fg_color=COLORS["section_frame"])
        search_row.pack(fill='x', padx=10, pady=5)

        self.edit_company_name_search_var = ctk.StringVar()
        self.edit_company_email_search_var = ctk.StringVar()

        ctk.CTkLabel(search_row, text="Company:", font=("Arial", 11, "bold")).pack(side='left', padx=3)
        name_entry = ctk.CTkEntry(search_row, textvariable=self.edit_company_name_search_var, width=180, height=32,
                                  placeholder_text="Company name")
        name_entry.pack(side='left', padx=3)
        self.edit_company_name_search_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(search_row, text="Email:", font=("Arial", 11, "bold")).pack(side='left', padx=3)
        email_entry = ctk.CTkEntry(search_row, textvariable=self.edit_company_email_search_var, width=180, height=32,
                                   placeholder_text="Email (optional)")
        email_entry.pack(side='left', padx=3)

        ctk.CTkButton(search_row, text="🔍 SEARCH", command=self.search_company_for_edit,
                      fg_color=COLORS["info"], font=("Arial", 11, "bold"), height=32, width=100).pack(side='left', padx=5)

        # Results/Edit form frame
        self.edit_company_form_frame = ctk.CTkFrame(tab, fg_color=COLORS["content_frame"])
        self.edit_company_form_frame.pack(fill='both', expand=True, padx=5, pady=3)

    def search_company_for_edit(self):
        company_name = self.edit_company_name_search_var.get().strip().upper()
        email = self.edit_company_email_search_var.get().strip()

        if not company_name and not email:
            messagebox.showerror("Error", "Please enter company name or email")
            return

        # Clear previous results
        for widget in self.edit_company_form_frame.winfo_children():
            widget.destroy()

        try:
            if self.collection is not None:
                # Build query based on inputs
                if company_name and email:
                    # Both fields - exact match
                    companies = list(self.collection.find({
                        "company_name": {"$regex": company_name, "$options": "i"},
                        "email": {"$regex": email, "$options": "i"}
                    }))
                elif company_name:
                    # Only company name - show all matching
                    companies = list(self.collection.find({
                        "company_name": {"$regex": company_name, "$options": "i"}
                    }))
                else:
                    # Only email
                    companies = list(self.collection.find({
                        "email": {"$regex": email, "$options": "i"}
                    }))

                if not companies:
                    ctk.CTkLabel(self.edit_company_form_frame, text="❌ No companies found",
                                 font=("Arial", 12), text_color=COLORS["error"]).pack(pady=20)
                    return

                if len(companies) == 1:
                    # Single result - show edit form directly
                    self.show_edit_company_form(companies[0])
                else:
                    # Multiple results - show list to select
                    ctk.CTkLabel(self.edit_company_form_frame, 
                                 text=f"📋 Found {len(companies)} companies - Click to edit:",
                                 font=("Arial", 12, "bold"), text_color=COLORS["info"]).pack(pady=5)

                    for company in companies:
                        row = ctk.CTkFrame(self.edit_company_form_frame, fg_color=COLORS["section_frame"], corner_radius=8)
                        row.pack(fill='x', padx=5, pady=2)

                        info = f"🏢 {company['company_name']}  |  📧 {company['email']}  |  👔 {company['hr_name']}  |  💰 {company['package']}"
                        ctk.CTkLabel(row, text=info, font=("Arial", 10)).pack(side='left', padx=10, pady=5)

                        ctk.CTkButton(row, text="✏️ EDIT", command=lambda c=company: self.show_edit_company_form(c),
                                      fg_color=COLORS["warning"], font=("Arial", 10, "bold"), 
                                      height=28, width=70).pack(side='right', padx=5, pady=3)
            else:
                messagebox.showerror("Error", "Database connection failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to search company: {e}")

    def show_edit_company_form(self, company):
        # Clear previous form
        for widget in self.edit_company_form_frame.winfo_children():
            widget.destroy()

        self.current_edit_company_id = company["_id"]

        # Create compact edit form
        company_frame = ctk.CTkFrame(self.edit_company_form_frame, fg_color=COLORS["section_frame"], corner_radius=10)
        company_frame.pack(fill='x', padx=5, pady=5)

        # Header with title and update button
        header_row = ctk.CTkFrame(company_frame, fg_color=COLORS["section_frame"])
        header_row.pack(fill='x', padx=10, pady=5)

        ctk.CTkLabel(header_row, text="📝 EDIT COMPANY INFORMATION",
                     font=("Arial", 14, "bold"), text_color=COLORS["warning"]).pack(side='left', padx=5)

        ctk.CTkButton(header_row, text="💾 UPDATE", command=self.update_company,
                      fg_color=COLORS["warning"], font=("Arial", 11, "bold"), height=32, width=100).pack(side='right', padx=5)

        # Initialize variables
        self.edit_company_name_var = ctk.StringVar(value=company["company_name"])
        self.edit_company_email_var = ctk.StringVar(value=company["email"])
        self.edit_contact_info_var = ctk.StringVar(value=company["contact_info"])
        self.edit_hr_name_var = ctk.StringVar(value=company["hr_name"])
        self.edit_package_var = ctk.StringVar(value=company["package"])
        self.edit_website_var = ctk.StringVar(value=company.get("website", ""))
        self.edit_address_var = ctk.StringVar(value=company.get("address", ""))

        # Compact 2-column grid layout
        company_grid = ctk.CTkFrame(company_frame, fg_color=COLORS["section_frame"])
        company_grid.pack(fill='x', padx=10, pady=5)

        # Row 1: Company Name, Email
        ctk.CTkLabel(company_grid, text="Company*:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', padx=3, pady=4)
        name_entry = ctk.CTkEntry(company_grid, textvariable=self.edit_company_name_var, width=200, height=32)
        name_entry.grid(row=0, column=1, sticky='w', padx=3, pady=4)
        self.edit_company_name_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(company_grid, text="Email*:", font=("Arial", 11, "bold")).grid(row=0, column=2, sticky='w', padx=3, pady=4)
        ctk.CTkEntry(company_grid, textvariable=self.edit_company_email_var, width=200, height=32).grid(row=0, column=3, sticky='w', padx=3, pady=4)

        # Row 2: Contact, HR Name
        ctk.CTkLabel(company_grid, text="Contact*:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky='w', padx=3, pady=4)
        ctk.CTkEntry(company_grid, textvariable=self.edit_contact_info_var, width=200, height=32).grid(row=1, column=1, sticky='w', padx=3, pady=4)

        ctk.CTkLabel(company_grid, text="HR Name*:", font=("Arial", 11, "bold")).grid(row=1, column=2, sticky='w', padx=3, pady=4)
        hr_entry = ctk.CTkEntry(company_grid, textvariable=self.edit_hr_name_var, width=200, height=32)
        hr_entry.grid(row=1, column=3, sticky='w', padx=3, pady=4)
        self.edit_hr_name_var.trace_add('write', self.uppercase_entry(hr_entry))

        # Row 3: Package, Website
        ctk.CTkLabel(company_grid, text="Package*:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', padx=3, pady=4)
        package_entry = ctk.CTkEntry(company_grid, textvariable=self.edit_package_var, width=200, height=32)
        package_entry.grid(row=2, column=1, sticky='w', padx=3, pady=4)
        self.edit_package_var.trace_add('write', self.uppercase_entry(package_entry))

        ctk.CTkLabel(company_grid, text="Website:", font=("Arial", 11, "bold")).grid(row=2, column=2, sticky='w', padx=3, pady=4)
        ctk.CTkEntry(company_grid, textvariable=self.edit_website_var, width=200, height=32).grid(row=2, column=3, sticky='w', padx=3, pady=4)

        # Row 4: Address
        ctk.CTkLabel(company_grid, text="Address:", font=("Arial", 11, "bold")).grid(row=3, column=0, sticky='w', padx=3, pady=4)
        address_entry = ctk.CTkEntry(company_grid, textvariable=self.edit_address_var, width=450, height=32)
        address_entry.grid(row=3, column=1, columnspan=3, sticky='w', padx=3, pady=4)
        self.edit_address_var.trace_add('write', self.uppercase_entry(address_entry))

    def update_company(self):
        # Validate required fields
        if not all([self.edit_company_name_var.get(), self.edit_company_email_var.get(),
                    self.edit_contact_info_var.get(), self.edit_hr_name_var.get(), self.edit_package_var.get()]):
            messagebox.showerror("Error", "Please fill all required fields (*)")
            return

        if not validate_email(self.edit_company_email_var.get()):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        # Prepare updated company data
        updated_data = {
            "company_name": self.edit_company_name_var.get().upper(),
            "email": self.edit_company_email_var.get(),
            "contact_info": self.edit_contact_info_var.get(),
            "hr_name": self.edit_hr_name_var.get().upper(),
            "package": self.edit_package_var.get().upper(),
            "website": self.edit_website_var.get(),
            "address": self.edit_address_var.get().upper()
        }

        try:
            if self.collection is not None:
                self.collection.update_one(
                    {"_id": self.current_edit_company_id},
                    {"$set": updated_data}
                )
                messagebox.showinfo("Success", "Company updated successfully!")
                # Clear edit form
                for widget in self.edit_company_form_frame.winfo_children():
                    widget.destroy()
                self.edit_company_search_var.set("")
            else:
                messagebox.showerror("Error", "Database connection failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update company: {e}")

    def setup_delete_company_tab(self, tab):
        # Title
        title_label = ctk.CTkLabel(tab, text="🗑️ DELETE COMPANY",
                                   font=("Arial", 28, "bold"), text_color=COLORS["error"])
        title_label.pack(pady=20)

        scroll_frame = ctk.CTkFrame(tab, fg_color=COLORS["content_frame"])
        scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Search frame
        delete_frame = ctk.CTkFrame(scroll_frame, fg_color=COLORS["section_frame"], corner_radius=15)
        delete_frame.pack(fill='x', padx=5, pady=8)

        ctk.CTkLabel(delete_frame, text="🔍 FIND COMPANY TO DELETE",
                     font=("Arial", 18, "bold"), text_color=COLORS["warning"]).pack(pady=10)

        search_controls = ctk.CTkFrame(delete_frame, fg_color=COLORS["section_frame"])
        search_controls.pack(fill='x', padx=20, pady=10)
        search_controls.columnconfigure(1, weight=1)

        self.delete_company_name_var = ctk.StringVar()
        self.delete_company_contact_var = ctk.StringVar()

        ctk.CTkLabel(search_controls, text="Company Name:", font=("Arial", 12, "bold")).grid(row=0, column=0,
                                                                                             sticky='w', padx=5,
                                                                                             pady=5)
        name_entry = ctk.CTkEntry(search_controls, textvariable=self.delete_company_name_var, width=300, height=35,
                                  font=("Arial", 12), placeholder_text="Enter company name")
        name_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        self.delete_company_name_var.trace_add('write', self.uppercase_entry(name_entry))

        ctk.CTkLabel(search_controls, text="Email/Contact:", font=("Arial", 12, "bold")).grid(row=1, column=0,
                                                                                              sticky='w', padx=5,
                                                                                              pady=5)
        contact_entry = ctk.CTkEntry(search_controls, textvariable=self.delete_company_contact_var, width=300,
                                     height=35, font=("Arial", 12), placeholder_text="Enter email or contact info")
        contact_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)
        # Contact can be email - no uppercase (email should stay as-is)

        search_btn = ctk.CTkButton(search_controls, text="🔍 SEARCH",
                                   command=self.search_company_for_delete,
                                   fg_color=COLORS["info"], font=("Arial", 13, "bold"), height=40, width=150)
        search_btn.grid(row=2, column=0, columnspan=2, pady=20)

        # Results frame
        self.delete_company_results_frame = ctk.CTkScrollableFrame(scroll_frame, fg_color=COLORS["content_frame"])
        self.delete_company_results_frame.pack(fill='both', expand=True, padx=10, pady=15)

    def search_company_for_delete(self):
        """Search for company by name and/or contact info"""
        name = self.delete_company_name_var.get().strip().upper()
        contact = self.delete_company_contact_var.get().strip().upper()

        if not name and not contact:
            messagebox.showerror("Error", "Please enter at least Company Name or Contact Info")
            return

        try:
            if self.collection is not None:
                # Build query
                query_conditions = []
                if name:
                    query_conditions.append({"company_name": {"$regex": name, "$options": "i"}})
                if contact:
                    query_conditions.append({
                        "$or": [
                            {"email": contact},
                            {"contact_info": contact}
                        ]
                    })

                if len(query_conditions) == 2:
                    query = {"$and": query_conditions}
                else:
                    query = query_conditions[0]

                companies = list(self.collection.find(query))

                # Clear previous results efficiently
                batch_clear_widgets(self.delete_company_results_frame)

                if not companies:
                    ctk.CTkLabel(self.delete_company_results_frame, text="❌ No companies found matching criteria",
                                 font=("Arial", 14), text_color=COLORS["error"]).pack(pady=20)
                    return

                # Display found companies
                for company in companies:
                    company_card = ctk.CTkFrame(self.delete_company_results_frame, fg_color=COLORS["section_frame"],
                                                corner_radius=10)
                    company_card.pack(fill='x', padx=5, pady=5)

                    info_text = (
                        f"Company: {company['company_name']}\n"
                        f"Email: {company['email']}\n"
                        f"Contact: {company['contact_info']}\n"
                        f"HR Name: {company['hr_name']}\n"
                        f"Package: {company['package']}"
                    )

                    ctk.CTkLabel(company_card, text=info_text, font=("Arial", 12), justify="left").pack(padx=20,
                                                                                                        pady=15)

                    delete_btn = ctk.CTkButton(company_card, text="🗑️ DELETE THIS COMPANY",
                                               command=lambda c=company: self.confirm_delete_company(c),
                                               fg_color=COLORS["error"], font=("Arial", 12, "bold"), height=40)
                    delete_btn.pack(padx=20, pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Search failed: {e}")

    def confirm_delete_company(self, company):
        """Confirm and delete company"""
        if messagebox.askyesno("Confirm Delete",
                               f"Are you sure you want to delete:\n{company['company_name']}?\n\nThis action cannot be undone!"):
            try:
                if self.collection is not None:
                    self.collection.delete_one({"_id": company["_id"]})
                    messagebox.showinfo("Success", "Company deleted successfully!")
                    self.delete_company_name_var.set("")
                    self.delete_company_contact_var.set("")
                    # Refresh results efficiently
                    batch_clear_widgets(self.delete_company_results_frame)
                else:
                    messagebox.showerror("Error", "Database connection failed")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete company: {e}")

    def export_companies_excel(self):
        """Export companies to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime

            # Get companies from stored data
            if not hasattr(self, 'current_companies') or not self.current_companies:
                messagebox.showwarning("No Data", "No companies to export. Please search first.")
                return

            companies = self.current_companies

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Companies"

            # Add headers
            headers = ["Company Name", "Email", "Contact", "HR Name", "Package", "Address"]
            ws.append(headers)

            # Style header
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data
            for company in companies:
                ws.append([
                    company.get("company_name", company.get("company_Name", "")),
                    company.get("email", ""),
                    company.get("contact_info", company.get("contact_no", "")),
                    company.get("hr_name", ""),
                    company.get("package", ""),
                    company.get("address", "")
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30

            # Open Save As dialog
            from tkinter import filedialog
            default_filename = f"companies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Companies As"
            )
            
            if not filename:  # User cancelled
                return
                
            wb.save(filename)
            messagebox.showinfo("Success", f"Companies exported to:\n{filename}")
        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed. Install it using: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {e}")

    def export_companies_csv(self):
        """Export companies to CSV file"""
        try:
            import csv
            from datetime import datetime
            from tkinter import filedialog

            # Get companies from stored data
            if not hasattr(self, 'current_companies') or not self.current_companies:
                messagebox.showwarning("No Data", "No companies to export. Please search first.")
                return

            companies = self.current_companies

            # Open Save As dialog
            default_filename = f"companies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Companies As"
            )
            
            if not filename:  # User cancelled
                return

            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ["Company Name", "Email", "Contact", "HR Name", "Package", "Address"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for company in companies:
                    writer.writerow({
                        "Company Name": company.get("company_name", company.get("company_Name", "")),
                        "Email": company.get("email", ""),
                        "Contact": company.get("contact_info", company.get("contact_no", "")),
                        "HR Name": company.get("hr_name", ""),
                        "Package": company.get("package", ""),
                        "Address": company.get("address", "")
                    })

            messagebox.showinfo("Success", f"Companies exported to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to CSV: {e}")

    def export_company_charts_to_excel(self):
        """Export company chart data to Excel with charts"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            from datetime import datetime
            from tkinter import filedialog

            if not hasattr(self, 'current_chart_data') or not self.current_chart_data:
                messagebox.showwarning("No Data", "No chart data to export. Please load charts first.")
                return

            companies = self.current_chart_data.get('companies', [])
            if not companies:
                messagebox.showwarning("No Data", "No company data to export.")
                return

            # Open Save As dialog
            default_filename = f"company_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Save Company Analytics As"
            )

            if not filename:  # User cancelled
                return

            # Create workbook
            wb = openpyxl.Workbook()

            # Styles
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # ========== Sheet 1: Industry Type Distribution (Pie Chart) ==========
            ws1 = wb.active
            ws1.title = "Industry Distribution"
            ws1.append(["Industry Type", "Number of Companies"])

            industry_labels, industry_counts = self.get_industry_distribution(companies)
            for i in range(len(industry_labels)):
                ws1.append([industry_labels[i], industry_counts[i]])

            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 22

            # Create Pie Chart
            if len(industry_labels) > 0:
                pie_chart = PieChart()
                pie_chart.title = "Industry Type Distribution"
                data = Reference(ws1, min_col=2, min_row=1, max_row=len(industry_labels) + 1)
                cats = Reference(ws1, min_col=1, min_row=2, max_row=len(industry_labels) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(cats)
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.width = 15
                pie_chart.height = 10
                ws1.add_chart(pie_chart, "D2")

            # ========== Sheet 2: Company Package Distribution (Bar Chart) ==========
            ws2 = wb.create_sheet("Package Distribution")
            ws2.append(["Package Range", "Number of Companies"])

            package_labels, package_counts = self.get_company_package_distribution(companies)
            for i in range(len(package_labels)):
                ws2.append([package_labels[i], package_counts[i]])

            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 22

            # Create Bar Chart
            if len(package_labels) > 0:
                bar_chart = BarChart()
                bar_chart.title = "Company Package Distribution"
                bar_chart.type = "col"
                bar_chart.style = 10
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(package_labels) + 1)
                cats = Reference(ws2, min_col=1, min_row=2, max_row=len(package_labels) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(cats)
                bar_chart.dataLabels = DataLabelList()
                bar_chart.dataLabels.showVal = True
                bar_chart.width = 15
                bar_chart.height = 10
                ws2.add_chart(bar_chart, "D2")

            # ========== Sheet 3: Top 20 Companies by Package (Horizontal Bar Chart) ==========
            ws3 = wb.create_sheet("Top 20 Companies")
            ws3.append(["Company Name", "Package (LPA)"])

            company_names, company_packages = self.get_top_companies_by_package(companies)
            for i in range(len(company_names)):
                ws3.append([company_names[i], company_packages[i]])

            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws3.column_dimensions['A'].width = 35
            ws3.column_dimensions['B'].width = 18

            # Create Horizontal Bar Chart
            if len(company_names) > 0:
                hbar_chart = BarChart()
                hbar_chart.title = "Top 20 Companies by Package"
                hbar_chart.type = "bar"  # Horizontal bar
                hbar_chart.style = 10
                data = Reference(ws3, min_col=2, min_row=1, max_row=min(len(company_names) + 1, 21))
                cats = Reference(ws3, min_col=1, min_row=2, max_row=min(len(company_names) + 1, 21))
                hbar_chart.add_data(data, titles_from_data=True)
                hbar_chart.set_categories(cats)
                hbar_chart.dataLabels = DataLabelList()
                hbar_chart.dataLabels.showVal = True
                hbar_chart.width = 18
                hbar_chart.height = 12
                ws3.add_chart(hbar_chart, "D2")

            # Save file
            wb.save(filename)
            
            # Count charts added
            chart_count = 0
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                chart_count += len(ws._charts)
            
            messagebox.showinfo("Success", 
                f"Company analytics exported successfully to:\n{filename}\n\n"
                f"📊 Charts added: {chart_count}\n"
                f"📋 Sheets: {len(wb.sheetnames)}")

        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed. Install it using: pip install openpyxl")
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("Error", f"Failed to export to Excel: {e}\n\nDetails:\n{error_details[:500]}")
